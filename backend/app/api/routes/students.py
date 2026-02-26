"""Student API endpoints - Story 3.9, 5.1, 5.5, 6.5, 9.9."""

import io
import logging
import re
import uuid
from datetime import UTC, datetime

from fastapi import File, HTTPException, Query, UploadFile, status
from fastapi.responses import Response
from fastapi.routing import APIRouter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlmodel import func, select

from app import crud
from app.api.deps import AsyncSessionDep, SessionDep, require_role
from app.models import (
    Activity,
    Assignment,
    AssignmentActivity,
    AssignmentPublishStatus,
    AssignmentStatus,
    AssignmentStudent,
    Class,
    ClassStudent,
    Student,
    StudentCreate,
    StudentPublic,
    Teacher,
    User,
    UserRole,
)
from app.schemas.analytics import (
    PeriodType,
    StudentAnalyticsResponse,
    StudentProgressPeriod,
    StudentProgressResponse,
)
from app.schemas.assignment import (
    StudentAssignmentResponse,
    StudentCalendarAssignmentItem,
    StudentCalendarAssignmentsResponse,
)
from app.schemas.feedback import StudentBadgeCountsResponse
from app.schemas.student_import import (
    CredentialsDownloadRequest,
    ImportCredential,
    ImportExecutionResponse,
    ImportRowResult,
    ImportRowStatus,
    ImportValidationResponse,
)
from app.models import (
    SkillCategory,
    StudentSkillScore,
)
from app.schemas.skill import (
    SkillProfileItem,
    SkillTrendLine,
    SkillTrendPoint,
    StudentSkillProfileResponse,
    StudentSkillTrendsResponse,
)
from app.services import analytics_service, feedback_service
from app.services.skill_attribution_service import _ACTIVITY_TYPE_SKILL_SLUG
from app.services.book_service_v2 import get_book_service
from app.utils import (
    ensure_unique_username,
    generate_student_password,
    generate_username_from_fullname,
    parse_excel_file,
    validate_file_size,
)

router = APIRouter(prefix="/students", tags=["students"])
logger = logging.getLogger(__name__)


@router.get(
    "/me/assignments",
    response_model=list[StudentAssignmentResponse],
    summary="Get student's assignments",
)
async def get_student_assignments(
    *,
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.student),
    status_filter: str | None = Query(None, alias="status"),
) -> list[StudentAssignmentResponse]:
    """
    Get all assignments for authenticated student.

    Query parameters:
        status: Filter by assignment status (not_started, in_progress, completed)

    Returns:
        List of assignments with enriched data (book, activity, progress)
    """
    # Get Student record from authenticated user
    result = await session.execute(
        select(Student).where(Student.user_id == current_user.id)
    )
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found"
        )

    # Subquery to count activities per assignment from AssignmentActivity junction table
    activity_count_subq = (
        select(
            AssignmentActivity.assignment_id,
            func.count(AssignmentActivity.activity_id).label("activity_count")
        )
        .group_by(AssignmentActivity.assignment_id)
        .subquery()
    )

    # Build query: AssignmentStudent → Assignment → Activity + activity count
    # Only show published assignments to students (not scheduled or draft)
    # Use LEFT JOIN (outerjoin) to include Content Library assignments without activity_id
    query = (
        select(
            AssignmentStudent,
            Assignment,
            Activity,
            func.coalesce(activity_count_subq.c.activity_count, 1).label("activity_count")
        )
        .join(Assignment, AssignmentStudent.assignment_id == Assignment.id)
        .outerjoin(Activity, Assignment.activity_id == Activity.id)
        .outerjoin(
            activity_count_subq,
            activity_count_subq.c.assignment_id == Assignment.id
        )
        .where(AssignmentStudent.student_id == student.id)
        .where(Assignment.status == AssignmentPublishStatus.published)
    )

    # Apply optional status filter
    if status_filter:
        # Validate status filter
        valid_statuses = ["not_started", "in_progress", "completed"]
        if status_filter not in valid_statuses:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Invalid status filter. Must be one of: {valid_statuses}"
            )
        query = query.where(AssignmentStudent.status == status_filter)

    # Execute query
    result = await session.execute(query)
    rows = result.all()

    # Get BookService to fetch book data from DCS
    book_service = get_book_service()

    # Map results to StudentAssignmentResponse
    assignments = []
    for row in rows:
        assignment_student = row[0]
        assignment = row[1]
        activity = row[2]  # May be None for Content Library assignments
        activity_count = row[3]

        # Fetch book data from DCS
        book = await book_service.get_book(assignment.dcs_book_id)

        # Handle Content Library assignments (may not have book in DCS)
        if not book:
            if assignment.activity_content:
                # Content Library assignment - use placeholder book info
                book_id = assignment.dcs_book_id if assignment.dcs_book_id else 0
                book_title = "AI Generated"
                book_cover_url = None
            else:
                logger.warning(
                    f"Book {assignment.dcs_book_id} not found in DCS for assignment {assignment.id}"
                )
                continue
        else:
            book_id = book.id
            book_title = book.title or book.name
            book_cover_url = book.cover_url

        # Handle Content Library assignments (no activity record)
        if activity:
            activity_id = activity.id
            activity_title = activity.title
            activity_type = activity.activity_type
        else:
            # Content Library assignment - use assignment's embedded data
            activity_id = assignment.id  # Use assignment id as pseudo activity id
            activity_title = assignment.name
            activity_type = assignment.activity_type if assignment.activity_type else "ai_quiz"

        response = StudentAssignmentResponse(
            # Assignment fields
            assignment_id=assignment.id,
            assignment_name=assignment.name,
            instructions=assignment.instructions,
            due_date=assignment.due_date,
            time_limit_minutes=assignment.time_limit_minutes,
            created_at=assignment.created_at,

            # Book fields (from DCS via BookService)
            book_id=book_id,
            book_title=book_title,
            book_cover_url=book_cover_url,

            # Activity fields (first activity for display purposes)
            activity_id=activity_id,
            activity_title=activity_title,
            activity_type=activity_type,

            # Student-specific fields
            status=assignment_student.status,
            score=assignment_student.score,
            started_at=assignment_student.started_at,
            completed_at=assignment_student.completed_at,
            time_spent_minutes=assignment_student.time_spent_minutes or 0,

            # Multi-activity support (Story 8.3)
            activity_count=activity_count,
        )
        assignments.append(response)

    return assignments


@router.get(
    "/me/progress",
    response_model=StudentProgressResponse,
    summary="Get student's own progress data",
)
async def get_student_progress(
    *,
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.student),
    period: StudentProgressPeriod = Query(
        default="this_month",
        description="Time period for progress data"
    ),
) -> StudentProgressResponse:
    """
    Get comprehensive progress data for the authenticated student.

    This endpoint is student-facing and provides encouraging, achievement-focused
    analytics designed for student motivation.

    Query parameters:
        period: Time period for data filtering
            - 'this_week': Current week (Monday to today)
            - 'this_month': Current month (default)
            - 'all_time': All completed work

    Returns:
        Complete progress data including:
        - Summary stats (avg score, streak, improvement trend)
        - Score trend over time (for chart display)
        - Activity type breakdown with user-friendly labels
        - Recent assignments (last 5)
        - Achievements/badges earned
        - Study time statistics
        - Personalized improvement tips
    """
    # Get Student record from authenticated user
    result = await session.execute(
        select(Student).where(Student.user_id == current_user.id)
    )
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found"
        )

    # Call analytics service
    try:
        progress = await analytics_service.get_student_progress(
            student_id=student.id,
            period=period,
            session=session,
        )
        return progress
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )


@router.get(
    "/{student_id}/analytics",
    response_model=StudentAnalyticsResponse,
    summary="Get student performance analytics",
)
async def get_student_analytics(
    *,
    session: AsyncSessionDep,
    student_id: uuid.UUID,
    period: PeriodType = Query(default="30d", description="Time period for analytics"),
    current_user: User = require_role(UserRole.teacher),
) -> StudentAnalyticsResponse:
    """
    Get comprehensive performance analytics for a student.

    **Authorization:** Teacher must have access to student via their classes.

    Query parameters:
        period: Time period for data filtering
            - '7d': Last 7 days
            - '30d': Last 30 days (default)
            - '3m': Last 3 months
            - 'all': All time

    Returns:
        Complete analytics including:
        - Summary metrics (avg score, completion rate, streak)
        - Recent activity (last 10 assignments)
        - Performance trend over time
        - Activity type breakdown
        - Assignment status summary
        - Time analytics
    """
    # Get teacher record
    teacher_result = await session.execute(
        select(Teacher).where(Teacher.user_id == current_user.id)
    )
    teacher = teacher_result.scalar_one_or_none()

    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teacher record not found"
        )

    # Verify teacher has access to student via their classes
    # Check if student is in any of the teacher's classes
    access_check = await session.execute(
        select(ClassStudent.id)
        .join(Class, ClassStudent.class_id == Class.id)
        .where(
            Class.teacher_id == teacher.id,
            ClassStudent.student_id == student_id,
        )
        .limit(1)
    )
    has_access = access_check.scalar_one_or_none() is not None

    if not has_access:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have access to this student's data"
        )

    # Call analytics service
    try:
        analytics = await analytics_service.get_student_analytics(
            student_id=student_id,
            period=period,
            session=session,
        )
        return analytics
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=str(e)
        )


@router.get(
    "/me/badges",
    response_model=StudentBadgeCountsResponse,
    summary="Get my badge counts (Story 6.5)",
)
async def get_my_badges(
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.student),
) -> StudentBadgeCountsResponse:
    """
    Get badge counts for the current student.

    Returns:
        Badge counts grouped by type, total, and this month counts

    Raises:
        HTTPException(404): Student record not found
    """
    # Get student record for current user
    student_result = await session.execute(
        select(Student).where(Student.user_id == current_user.id)
    )
    student = student_result.scalar_one_or_none()

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found"
        )

    return await feedback_service.get_student_badge_counts(
        session=session,
        student_id=student.id,
    )


@router.get(
    "/{student_id}/badges",
    response_model=StudentBadgeCountsResponse,
    summary="Get student's badge counts (Story 6.5)",
)
async def get_student_badges(
    student_id: uuid.UUID,
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.teacher, UserRole.student),
) -> StudentBadgeCountsResponse:
    """
    Get badge counts for a student across all their feedback.

    Students can view their own badges. Teachers can view badges for
    students in their classes.

    Args:
        student_id: UUID of the student (students.id)
        session: Database session
        current_user: Authenticated user (teacher or student)

    Returns:
        Badge counts grouped by type, total, and this month counts

    Raises:
        HTTPException(403): Not authorized to view this student's badges
        HTTPException(404): Student not found
    """
    # Verify access
    if current_user.role == UserRole.student:
        # Student can only view their own badges
        student_result = await session.execute(
            select(Student).where(Student.user_id == current_user.id)
        )
        student = student_result.scalar_one_or_none()
        if not student or student.id != student_id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You can only view your own badges"
            )
    else:
        # Teacher must have student in their class
        teacher_result = await session.execute(
            select(Teacher).where(Teacher.user_id == current_user.id)
        )
        teacher = teacher_result.scalar_one_or_none()

        if not teacher:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teacher record not found"
            )

        access_check = await session.execute(
            select(ClassStudent.id)
            .join(Class, ClassStudent.class_id == Class.id)
            .where(
                Class.teacher_id == teacher.id,
                ClassStudent.student_id == student_id,
            )
            .limit(1)
        )
        has_access = access_check.scalar_one_or_none() is not None

        if not has_access:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this student's badges"
            )

    # Verify student exists
    student_check = await session.execute(
        select(Student).where(Student.id == student_id)
    )
    if not student_check.scalar_one_or_none():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found"
        )

    # Get badge counts
    badge_data = await feedback_service.get_student_badge_counts(
        db=session,
        student_id=student_id,
    )

    return StudentBadgeCountsResponse(**badge_data)


# --- Unassigned Students Endpoint (Story 20.5) ---


@router.get(
    "",
    response_model=list[StudentPublic],
    summary="Get all students (admin/supervisor) or teacher's students",
)
async def get_students(
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.admin, UserRole.supervisor, UserRole.teacher),
    limit: int = Query(100, ge=1, le=500, description="Maximum number of students to return"),
    offset: int = Query(0, ge=0, description="Number of students to skip"),
) -> list[StudentPublic]:
    """
    Get list of students based on user role.

    **Permissions:**
    - Admin/Supervisor: See all students in the system
    - Teacher: See only students in their classes

    Args:
        session: Database session
        current_user: Authenticated user
        limit: Maximum number of students to return (default 100, max 500)
        offset: Number of students to skip for pagination

    Returns:
        List of Student records with user information

    Raises:
        HTTPException(404): Teacher record not found (for teacher role)
    """
    if current_user.role in [UserRole.admin, UserRole.supervisor]:
        # Admin and Supervisor can see all students
        query = (
            select(Student, User)
            .join(User, Student.user_id == User.id)
            .offset(offset)
            .limit(limit)
        )
    elif current_user.role == UserRole.teacher:
        # Teachers can only see students in their classes
        teacher_result = await session.execute(
            select(Teacher).where(Teacher.user_id == current_user.id)
        )
        teacher = teacher_result.scalar_one_or_none()

        if not teacher:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teacher record not found"
            )

        # Get students from teacher's classes
        query = (
            select(Student, User)
            .join(User, Student.user_id == User.id)
            .join(ClassStudent, ClassStudent.student_id == Student.id)
            .join(Class, Class.id == ClassStudent.class_id)
            .where(Class.teacher_id == teacher.id)
            .distinct()
            .offset(offset)
            .limit(limit)
        )
    else:
        # Should not reach here due to require_role, but just in case
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You do not have permission to view students"
        )

    result = await session.execute(query)
    rows = result.all()

    # Build StudentPublic response
    students = []
    for student, user in rows:
        students.append(
            StudentPublic(
                id=student.id,
                user_id=user.id,
                user_email=user.email,
                user_username=user.username,
                user_full_name=user.full_name,
                grade_level=student.grade_level,
                parent_email=student.parent_email,
                created_at=student.created_at,
                updated_at=student.updated_at,
            )
        )

    return students


@router.get(
    "/unassigned",
    response_model=list[StudentPublic],
    summary="Get students not enrolled in teacher's classes (Story 20.5)",
)
async def get_unassigned_students(
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.teacher),
) -> list[StudentPublic]:
    """
    Get students visible to teacher but not in any of teacher's classes.

    This endpoint returns students who:
    - Were created by the current teacher (via created_by_teacher_id)
    - Are NOT enrolled in any of the teacher's classes

    This allows teachers to assign individual students who aren't in classes yet.

    Args:
        session: Database session
        current_user: Authenticated teacher user

    Returns:
        List of Student records with user information

    Raises:
        HTTPException(404): Teacher record not found
    """
    # Get teacher record
    teacher_result = await session.execute(
        select(Teacher).where(Teacher.user_id == current_user.id)
    )
    teacher = teacher_result.scalar_one_or_none()

    if not teacher:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Teacher record not found"
        )

    # Get all student IDs that are enrolled in this teacher's classes
    enrolled_student_ids_result = await session.execute(
        select(ClassStudent.student_id)
        .join(Class, ClassStudent.class_id == Class.id)
        .where(Class.teacher_id == teacher.id)
    )
    enrolled_student_ids = {row[0] for row in enrolled_student_ids_result.all()}

    # Get all students created by this teacher who are NOT enrolled in any of their classes
    unassigned_result = await session.execute(
        select(Student, User)
        .join(User, Student.user_id == User.id)
        .where(
            Student.created_by_teacher_id == teacher.id,
            ~Student.id.in_(enrolled_student_ids) if enrolled_student_ids else True,
        )
    )
    rows = unassigned_result.all()

    # Build StudentPublic response
    unassigned_students = []
    for student, user in rows:
        unassigned_students.append(
            StudentPublic(
                id=student.id,
                user_id=user.id,
                user_email=user.email,
                user_username=user.username,
                user_full_name=user.full_name,
                grade_level=student.grade_level,
                parent_email=student.parent_email,
                created_at=student.created_at,
                updated_at=student.updated_at,
            )
        )

    return unassigned_students


# --- Student Calendar Endpoints ---


@router.get(
    "/me/calendar",
    response_model=StudentCalendarAssignmentsResponse,
    summary="Get student's assignments for calendar view",
    description="Get assignments within a date range for calendar display. Only shows published assignments.",
)
async def get_student_calendar_assignments(
    *,
    session: AsyncSessionDep,
    start_date: datetime = Query(..., description="Start date for range (inclusive)"),
    end_date: datetime = Query(..., description="End date for range (inclusive)"),
    current_user: User = require_role(UserRole.student),
) -> StudentCalendarAssignmentsResponse:
    """
    Get assignments for student calendar view.

    **Query Parameters:**
    - start_date: Start of date range (required)
    - end_date: End of date range (required)

    **Returns:**
    Assignments grouped by due date within the specified range.
    Only shows published assignments (scheduled ones are not visible until published).
    """
    from collections import defaultdict

    # Get Student record for current user
    result = await session.execute(
        select(Student).where(Student.user_id == current_user.id)
    )
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found for this user",
        )

    # Subquery to count activities per assignment
    activity_count_subq = (
        select(
            AssignmentActivity.assignment_id,
            func.count(AssignmentActivity.activity_id).label("activity_count")
        )
        .group_by(AssignmentActivity.assignment_id)
        .subquery()
    )

    # Query published assignments in date range
    # Show assignments based on due_date, scheduled_publish_date, OR created_at (fallback)
    from sqlalchemy import or_

    query = (
        select(
            AssignmentStudent,
            Assignment,
            func.coalesce(activity_count_subq.c.activity_count, 1).label("activity_count")
        )
        .join(Assignment, AssignmentStudent.assignment_id == Assignment.id)
        .outerjoin(
            activity_count_subq,
            activity_count_subq.c.assignment_id == Assignment.id
        )
        .where(AssignmentStudent.student_id == student.id)
        .where(Assignment.status == AssignmentPublishStatus.published)
        .where(
            # Include if due_date, scheduled_publish_date, OR created_at falls in range
            or_(
                (Assignment.due_date >= start_date) & (Assignment.due_date <= end_date),
                (Assignment.scheduled_publish_date >= start_date) & (Assignment.scheduled_publish_date <= end_date),
                (Assignment.created_at >= start_date) & (Assignment.created_at <= end_date)
            )
        )
    )

    result = await session.execute(query)
    rows = result.all()

    # Get BookService to fetch book data from DCS
    book_service = get_book_service()

    # Group assignments by date (priority: due_date > scheduled_publish_date > created_at)
    assignments_by_date: dict[str, list[StudentCalendarAssignmentItem]] = defaultdict(list)

    for row in rows:
        assignment_student = row[0]
        assignment = row[1]
        activity_count = row[2]

        # Fetch book data from DCS
        book = await book_service.get_book(assignment.dcs_book_id)
        if not book:
            logger.warning(
                f"Book {assignment.dcs_book_id} not found in DCS for assignment {assignment.id}"
            )
            continue

        # Use due_date as primary, scheduled_publish_date as secondary, created_at as fallback
        calendar_date = assignment.due_date or assignment.scheduled_publish_date or assignment.created_at
        date_key = calendar_date.strftime("%Y-%m-%d")

        calendar_item = StudentCalendarAssignmentItem(
            id=assignment.id,
            name=assignment.name,
            due_date=assignment.due_date,
            book_id=book.id,
            book_title=book.title or book.name,
            book_cover_url=book.cover_url,
            activity_count=activity_count,
            status=assignment_student.status.value if hasattr(assignment_student.status, 'value') else str(assignment_student.status),
        )
        assignments_by_date[date_key].append(calendar_item)

    # Sort assignments within each day
    for date_key in assignments_by_date:
        assignments_by_date[date_key].sort(key=lambda a: a.due_date or datetime.min.replace(tzinfo=UTC))

    total_assignments = sum(len(items) for items in assignments_by_date.values())

    logger.info(
        f"Student calendar query: student_id={student.id}, "
        f"range={start_date.date()}-{end_date.date()}, "
        f"assignments={total_assignments}"
    )

    return StudentCalendarAssignmentsResponse(
        start_date=start_date.strftime("%Y-%m-%d"),
        end_date=end_date.strftime("%Y-%m-%d"),
        total_assignments=total_assignments,
        assignments_by_date=dict(assignments_by_date),
    )


# --- Student Import Endpoints (Story 9.9) ---

# Template column headers (Story 28.1: Added Password column)
IMPORT_TEMPLATE_HEADERS = [
    "Full Name *",
    "Username",
    "Password",  # Story 28.1: Optional - use class_password or auto-generate if empty
    "Email",
    "Parent Email",
    "Grade",
    "Class",
]


def _create_import_template_workbook() -> Workbook:
    """Create the student import template Excel workbook."""
    wb = Workbook()

    # --- Students Sheet ---
    ws_students = wb.active
    ws_students.title = "Students"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col, header in enumerate(IMPORT_TEMPLATE_HEADERS, start=1):
        cell = ws_students.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Add example row (Story 28.1: Added password column)
    example_row = [
        "Neşet Ertaş",  # Full Name
        "neset.ertas",  # Username (optional)
        "student123",  # Password (optional - leave empty to auto-generate)
        "neset@email.com",  # Email
        "parent@email.com",  # Parent Email
        "3",  # Grade
        "A",  # Class
    ]
    for col, value in enumerate(example_row, start=1):
        ws_students.cell(row=2, column=col, value=value)

    # Set column widths
    column_widths = [20, 20, 15, 25, 25, 15, 15]
    for col, width in enumerate(column_widths, start=1):
        ws_students.column_dimensions[get_column_letter(col)].width = width

    # --- Instructions Sheet ---
    ws_instructions = wb.create_sheet(title="Instructions")

    instructions = [
        ("Student Import Template Instructions", None),
        ("", None),
        ("Column Descriptions:", None),
        ("- Full Name * (Required):", "Student's full name. Used to generate username if not provided."),
        ("- Username (Optional):", "Leave empty to auto-generate from full name. Turkish characters will be converted."),
        ("- Password (Optional):", "Student's password. Leave empty to auto-generate or use class password option."),
        ("- Email (Optional):", "Student's email address."),
        ("- Parent Email (Optional):", "Parent/guardian email address for communications."),
        ("- Grade (Optional):", "Student's grade level (e.g., '3', '5', '10')."),
        ("- Class (Optional):", "Class/section identifier (e.g., 'A', 'B', 'Morning'). Used with Grade to create/assign classrooms."),
        ("", None),
        ("Password Options:", None),
        ("- Provide passwords in the Password column for individual students", None),
        ("- Leave Password column empty and use 'class password' option to set same password for all", None),
        ("- Leave Password column empty without class password to auto-generate unique passwords", None),
        ("", None),
        ("Notes:", None),
        ("- Maximum 500 students per upload", None),
        ("- Usernames are auto-generated as: firstname.lastname (lowercase, Turkish chars converted)", None),
        ("- If username exists, a number will be appended (e.g., john.doe2)", None),
        ("- Auto-generated passwords are 8 characters with letters and numbers", None),
        ("- Teachers can view and change student passwords anytime from the student list", None),
        ("- Students are bound to the teacher who imports them", None),
        ("", None),
        ("Turkish Character Conversion:", None),
        ("ı → i, İ → I, ğ → g, Ğ → G, ü → u, Ü → U, ş → s, Ş → S, ö → o, Ö → O, ç → c, Ç → C", None),
    ]

    for row_num, (text, description) in enumerate(instructions, start=1):
        cell = ws_instructions.cell(row=row_num, column=1, value=text)
        if row_num == 1:
            cell.font = Font(bold=True, size=14)
        elif text.startswith("-"):
            cell.font = Font(bold=True)
        if description:
            ws_instructions.cell(row=row_num, column=2, value=description)

    ws_instructions.column_dimensions["A"].width = 50
    ws_instructions.column_dimensions["B"].width = 60

    return wb


@router.get(
    "/import-template",
    summary="Download student import template (Story 9.9)",
    description="Download Excel template for bulk student import with headers and instructions.",
    responses={
        200: {
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            },
            "description": "Excel template file",
        }
    },
)
def get_import_template(
    _current_user: User = require_role(UserRole.admin, UserRole.teacher),
) -> Response:
    """
    Generate and return the student import template Excel file.

    **Permissions:** Admin or Teacher

    Returns an Excel file with:
    - Students sheet with column headers and example rows
    - Instructions sheet explaining each column

    AC: 2, 3, 4, 5, 6, 7, 8
    """
    wb = _create_import_template_workbook()

    # Save to bytes buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=student_import_template.xlsx",
            "Content-Length": str(len(content)),
            "Cache-Control": "no-cache, no-store, must-revalidate",
        },
    )


def _validate_email_format(email: str) -> bool:
    """Validate email format using regex."""
    if not email:
        return True  # Empty email is valid (optional field)
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


@router.post(
    "/import/validate",
    response_model=ImportValidationResponse,
    summary="Validate student import file (Story 9.9)",
    description="Upload and validate Excel file for student import. Returns validation results for each row.",
)
async def validate_import_file(
    session: SessionDep,
    file: UploadFile = File(..., description="Excel file (.xlsx or .xls)"),
    _current_user: User = require_role(UserRole.admin, UserRole.teacher),
) -> ImportValidationResponse:
    """
    Validate an uploaded Excel file for student import.

    **Permissions:** Admin or Teacher

    Validates:
    - File type (.xlsx, .xls)
    - File size (max 5MB)
    - Row count (max 500)
    - Required fields (Full Name)
    - Email format if provided
    - Username uniqueness in file and database

    AC: 11, 12, 13, 14, 15, 16, 17, 18, 19, 20
    """
    # Validate file type
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="No filename provided"
        )

    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are accepted"
        )

    # Validate file size
    is_valid_size = await validate_file_size(file, max_size_mb=5)
    if not is_valid_size:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds 5MB limit"
        )

    # Parse Excel file
    try:
        rows = await parse_excel_file(file)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

    # Validate row count
    if len(rows) > 500:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Maximum 500 students per upload. File contains {len(rows)} rows."
        )

    if len(rows) == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File contains no data rows"
        )

    # Get existing usernames from database for uniqueness check
    existing_usernames_result = session.execute(
        select(User.username).where(User.username.isnot(None))
    )
    existing_usernames: set[str] = {row[0] for row in existing_usernames_result.all() if row[0]}

    # Track usernames seen in this file
    file_usernames: set[str] = set()

    # Validate each row
    results: list[ImportRowResult] = []
    valid_count = 0
    warning_count = 0
    error_count = 0

    for row in rows:
        row_num = row.get("_row_number", 0)
        # Support both new ("Full Name *") and old ("Full Name") column headers
        full_name = str(row.get("Full Name *", "") or row.get("Full Name", "") or "").strip()
        provided_username = str(row.get("Username", "") or "").strip()
        email = str(row.get("Email", "") or "").strip()
        parent_email = str(row.get("Parent Email", "") or "").strip()
        grade = str(row.get("Grade", "") or "").strip()
        class_name = str(row.get("Class", "") or "").strip()

        errors: list[str] = []
        warnings: list[str] = []

        # Required field validation
        if not full_name:
            errors.append("Full Name is required")

        # Email format validation
        if email and not _validate_email_format(email):
            errors.append("Invalid email format")

        # Parent email format validation
        if parent_email and not _validate_email_format(parent_email):
            errors.append("Invalid parent email format")

        # Generate or validate username
        if provided_username:
            username = provided_username.lower()
        elif full_name:
            username = generate_username_from_fullname(full_name)
        else:
            username = ""

        # Check username uniqueness
        if username:
            # Check against database
            if username in existing_usernames:
                # Auto-fix: will append number
                original = username
                username = ensure_unique_username(username, existing_usernames | file_usernames)
                warnings.append(f"Username '{original}' exists, will use '{username}'")

            # Check against file (duplicates within file)
            if username in file_usernames:
                original = username
                username = ensure_unique_username(username, existing_usernames | file_usernames)
                warnings.append(f"Duplicate username in file, will use '{username}'")

            file_usernames.add(username)

        # Determine status
        if errors:
            row_status = ImportRowStatus.error
            error_count += 1
        elif warnings:
            row_status = ImportRowStatus.warning
            warning_count += 1
        else:
            row_status = ImportRowStatus.valid
            valid_count += 1

        results.append(ImportRowResult(
            row_number=row_num,
            full_name=full_name,
            username=username,
            email=email if email else None,
            grade=grade if grade else None,
            class_name=class_name if class_name else None,
            status=row_status,
            errors=errors,
            warnings=warnings,
        ))

    return ImportValidationResponse(
        valid_count=valid_count,
        warning_count=warning_count,
        error_count=error_count,
        total_count=len(rows),
        rows=results,
    )


@router.post(
    "/import",
    response_model=ImportExecutionResponse,
    summary="Execute student import (Story 9.9, 28.1)",
    description="Import students from validated Excel file. Returns created credentials.",
)
async def execute_import(
    session: SessionDep,
    file: UploadFile = File(..., description="Excel file (.xlsx or .xls)"),
    school_id: uuid.UUID | None = Query(None, description="School ID (required for Admin)"),
    teacher_id_param: uuid.UUID | None = Query(None, alias="teacher_id", description="Teacher ID (optional for Admin, to assign classrooms)"),
    class_password: str | None = Query(None, min_length=4, max_length=50, description="Password to apply to all students (Story 28.1)"),
    current_user: User = require_role(UserRole.admin, UserRole.teacher),
) -> ImportExecutionResponse:
    """
    Execute student import from Excel file.

    **Permissions:** Admin or Teacher

    - Admin must provide school_id, optionally teacher_id for classroom creation
    - Teacher uses their own school

    **Password Options (Story 28.1):**
    - Password from Excel file (Password column) takes priority
    - If no password in file, class_password is used (same for all)
    - If neither, auto-generate unique password for each student

    Creates students with:
    - Role: student
    - Generated/provided username and password
    - Association with school

    AC: 24, 25, 26, 27, 30, 31, 32
    """
    # Determine school_id based on role
    teacher_id: uuid.UUID | None = None

    if current_user.role == UserRole.teacher:
        # Teacher imports to their own school
        teacher_result = session.execute(
            select(Teacher).where(Teacher.user_id == current_user.id)
        )
        teacher = teacher_result.scalar_one_or_none()
        if not teacher:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teacher record not found"
            )
        if not teacher.school_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Teacher must be associated with a school to import students"
            )
        school_id = teacher.school_id
        teacher_id = teacher.id
    elif current_user.role == UserRole.admin:
        if not school_id:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Admin must specify school_id for import"
            )
        # Admin can optionally specify a teacher for classroom creation
        if teacher_id_param:
            # Verify teacher exists and belongs to the selected school
            teacher_result = session.execute(
                select(Teacher).where(
                    Teacher.id == teacher_id_param,
                    Teacher.school_id == school_id
                )
            )
            teacher = teacher_result.scalar_one_or_none()
            if not teacher:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Teacher not found or does not belong to selected school"
                )
            teacher_id = teacher.id

    # Validate file
    if not file.filename or not file.filename.lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are accepted"
        )

    is_valid_size = await validate_file_size(file, max_size_mb=5)
    if not is_valid_size:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="File size exceeds 5MB limit"
        )

    # Parse Excel file
    try:
        rows = await parse_excel_file(file)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=str(e)
        )

    if len(rows) > 500:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Maximum 500 students per upload. File contains {len(rows)} rows."
        )

    # Get existing usernames
    existing_usernames_result = session.execute(
        select(User.username).where(User.username.isnot(None))
    )
    existing_usernames: set[str] = {row[0] for row in existing_usernames_result.all() if row[0]}

    # Track usernames used in this import
    used_usernames: set[str] = set(existing_usernames)

    # Cache for classrooms: (grade, class_name) -> Class object
    classroom_cache: dict[tuple[str, str], Class] = {}

    def get_or_create_classroom(grade: str, class_name: str) -> Class | None:
        """Get existing classroom or create new one for the grade/class combination."""
        if not grade or not class_name:
            return None

        cache_key = (grade, class_name)
        if cache_key in classroom_cache:
            return classroom_cache[cache_key]

        # Build classroom name like "4B" or "Grade 4 - B"
        classroom_name = f"{grade}{class_name}"

        # Check if classroom already exists for this teacher/school
        existing_class = session.execute(
            select(Class).where(
                Class.name == classroom_name,
                Class.school_id == school_id,
                Class.teacher_id == teacher_id if teacher_id else True,
            )
        ).scalar_one_or_none()

        if existing_class:
            classroom_cache[cache_key] = existing_class
            return existing_class

        # Create new classroom
        new_class = Class(
            name=classroom_name,
            grade_level=grade,
            teacher_id=teacher_id,
            school_id=school_id,
            is_active=True,
        )
        session.add(new_class)
        session.flush()  # Get the ID
        classroom_cache[cache_key] = new_class
        logger.info(f"Auto-created classroom: {classroom_name} (Grade: {grade})")
        return new_class

    def enroll_student_in_classroom(student: Student, classroom: Class) -> None:
        """Enroll student in classroom if not already enrolled."""
        # Check if already enrolled
        existing_enrollment = session.execute(
            select(ClassStudent).where(
                ClassStudent.class_id == classroom.id,
                ClassStudent.student_id == student.id,
            )
        ).scalar_one_or_none()

        if not existing_enrollment:
            enrollment = ClassStudent(
                class_id=classroom.id,
                student_id=student.id,
            )
            session.add(enrollment)

    # Import students
    credentials: list[ImportCredential] = []
    errors: list[str] = []
    created_count = 0
    failed_count = 0
    classrooms_created = 0

    for row in rows:
        row_num = row.get("_row_number", 0)
        # Support both new ("Full Name *") and old ("Full Name") column headers
        full_name = str(row.get("Full Name *", "") or row.get("Full Name", "") or "").strip()
        provided_username = str(row.get("Username", "") or "").strip()
        row_password = str(row.get("Password", "") or "").strip()  # Story 28.1
        email = str(row.get("Email", "") or "").strip()
        parent_email = str(row.get("Parent Email", "") or "").strip()
        grade = str(row.get("Grade", "") or "").strip()
        class_name = str(row.get("Class", "") or "").strip()

        # Skip rows without full name
        if not full_name:
            errors.append(f"Row {row_num}: Skipped - Full Name is required")
            failed_count += 1
            continue

        try:
            # Generate username if not provided
            if provided_username:
                username = ensure_unique_username(
                    provided_username.lower(),
                    used_usernames
                )
            else:
                base_username = generate_username_from_fullname(full_name)
                username = ensure_unique_username(base_username, used_usernames)

            used_usernames.add(username)

            # Story 28.1: Password priority: row password > class password > auto-generate
            if row_password:
                password = row_password
            elif class_password:
                password = class_password
            else:
                password = generate_student_password()

            # Email is optional - pass None if not provided
            user_email = email if email else None

            # Build grade_level from Grade and Class columns
            grade_level = None
            if grade and class_name:
                grade_level = f"{grade}{class_name}"  # e.g., "3A"
            elif grade:
                grade_level = grade

            # Create student using crud function
            student_create = StudentCreate(
                user_id=uuid.uuid4(),  # Placeholder
                grade_level=grade_level,
                parent_email=parent_email if parent_email else None,
            )

            user, student = crud.create_student(
                session=session,
                email=user_email,
                username=username,
                password=password,
                full_name=full_name,
                student_create=student_create,
                created_by_teacher_id=teacher_id,
            )

            # Auto-create classroom and enroll student if grade and class provided
            if grade and class_name and teacher_id:
                cache_size_before = len(classroom_cache)
                classroom = get_or_create_classroom(grade, class_name)
                if classroom:
                    if len(classroom_cache) > cache_size_before:
                        classrooms_created += 1
                    enroll_student_in_classroom(student, classroom)

            credentials.append(ImportCredential(
                full_name=full_name,
                username=username,
                password=password,
                email=email if email else None,
            ))
            created_count += 1

        except Exception as e:
            logger.error(f"Failed to create student from row {row_num}: {e}")
            errors.append(f"Row {row_num}: {str(e)}")
            failed_count += 1

    # Commit all changes
    session.commit()

    # Add info about auto-created classrooms to response
    if classrooms_created > 0:
        errors.insert(0, f"Info: Auto-created {classrooms_created} classroom(s) based on Grade/Class data")

    return ImportExecutionResponse(
        created_count=created_count,
        failed_count=failed_count,
        credentials=credentials,
        errors=errors,
    )


@router.post(
    "/import/credentials",
    summary="Download credentials file (Story 9.9)",
    description="Generate Excel file with student credentials for download.",
    responses={
        200: {
            "content": {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}
            },
            "description": "Excel credentials file",
        }
    },
)
def download_credentials(
    request: CredentialsDownloadRequest,
    _current_user: User = require_role(UserRole.admin, UserRole.teacher),
) -> Response:
    """
    Generate and download credentials Excel file.

    **Permissions:** Admin or Teacher

    This is a one-time download - passwords are not stored
    and cannot be retrieved later.

    AC: 28, 29
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Credentials"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Headers
    headers = ["Full Name", "Username", "Password", "Email"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row_num, cred in enumerate(request.credentials, start=2):
        ws.cell(row=row_num, column=1, value=cred.full_name)
        ws.cell(row=row_num, column=2, value=cred.username)
        ws.cell(row=row_num, column=3, value=cred.password)
        ws.cell(row=row_num, column=4, value=cred.email or "")

    # Column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 30

    # Add warning row at top
    ws.insert_rows(1)
    warning_cell = ws.cell(row=1, column=1, value="WARNING: Store this file securely! Passwords cannot be retrieved later.")
    warning_cell.font = Font(bold=True, color="FF0000")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()

    timestamp = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")

    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=student_credentials_{timestamp}.xlsx"
        },
    )


# =============================================================================
# Student Skill Profile Endpoints (Story 30.14)
# =============================================================================


async def _build_skill_profile(
    student_id: uuid.UUID,
    student_name: str,
    session: AsyncSessionDep,
) -> StudentSkillProfileResponse:
    """Build skill profile for a student from StudentSkillScore data,
    with fallback to deriving from AssignmentStudent scores when no
    StudentSkillScore records exist."""
    from sqlalchemy import case, distinct

    # Count total AI assignments completed (assignments with skill scores)
    total_query = await session.execute(
        select(func.count(distinct(StudentSkillScore.assignment_id)))
        .where(StudentSkillScore.student_id == student_id)
    )
    total_ai_assignments = total_query.scalar() or 0

    # Get all active skills to show all 5 axes even if no data
    skills_result = await session.execute(
        select(SkillCategory).where(SkillCategory.is_active == True)
    )
    all_skills = skills_result.scalars().all()
    skill_by_slug: dict[str, SkillCategory] = {s.slug: s for s in all_skills}

    score_map: dict[uuid.UUID, dict] = {}
    trend_map: dict[uuid.UUID, str | None] = {}

    if total_ai_assignments > 0:
        # Primary path: use StudentSkillScore records
        scores_query = await session.execute(
            select(
                StudentSkillScore.skill_id,
                func.sum(StudentSkillScore.attributed_score).label("total_score"),
                func.sum(StudentSkillScore.attributed_max_score).label("total_max"),
                func.count().label("data_points"),
            )
            .where(StudentSkillScore.student_id == student_id)
            .group_by(StudentSkillScore.skill_id)
        )
        for row in scores_query.all():
            score_map[row.skill_id] = {
                "total_score": float(row.total_score),
                "total_max": float(row.total_max),
                "data_points": int(row.data_points),
            }

        # Build trend data for skills with enough data points
        for skill_id, data in score_map.items():
            if data["data_points"] >= 6:
                trend_query = await session.execute(
                    select(
                        StudentSkillScore.attributed_score,
                        StudentSkillScore.attributed_max_score,
                    )
                    .where(
                        StudentSkillScore.student_id == student_id,
                        StudentSkillScore.skill_id == skill_id,
                    )
                    .order_by(StudentSkillScore.recorded_at.desc())
                    .limit(6)
                )
                recent_scores = trend_query.all()
                if len(recent_scores) >= 6:
                    recent_3 = [
                        (s.attributed_score / s.attributed_max_score * 100)
                        if s.attributed_max_score > 0 else 0
                        for s in recent_scores[:3]
                    ]
                    prev_3 = [
                        (s.attributed_score / s.attributed_max_score * 100)
                        if s.attributed_max_score > 0 else 0
                        for s in recent_scores[3:6]
                    ]
                    diff = sum(recent_3) / 3 - sum(prev_3) / 3
                    if diff > 5:
                        trend_map[skill_id] = "improving"
                    elif diff < -5:
                        trend_map[skill_id] = "declining"
                    else:
                        trend_map[skill_id] = "stable"
                else:
                    trend_map[skill_id] = None
            else:
                trend_map[skill_id] = None
    else:
        # Fallback: derive from completed AI content assignments directly
        fallback_query = await session.execute(
            select(AssignmentStudent.score, Assignment.activity_type)
            .join(Assignment, AssignmentStudent.assignment_id == Assignment.id)
            .where(
                AssignmentStudent.student_id == student_id,
                AssignmentStudent.status == AssignmentStatus.completed,
                AssignmentStudent.score.isnot(None),
                Assignment.activity_type.isnot(None),
            )
        )
        fallback_rows = fallback_query.all()

        if fallback_rows:
            total_ai_assignments = len(fallback_rows)

            # Aggregate by inferred skill
            for row in fallback_rows:
                act_type_val = row.activity_type.value if hasattr(row.activity_type, 'value') else str(row.activity_type)
                skill_slug = _ACTIVITY_TYPE_SKILL_SLUG.get(act_type_val)
                if not skill_slug:
                    continue
                skill = skill_by_slug.get(skill_slug)
                if not skill:
                    continue

                if skill.id not in score_map:
                    score_map[skill.id] = {"total_score": 0.0, "total_max": 0.0, "data_points": 0}
                score_map[skill.id]["total_score"] += float(row.score)
                score_map[skill.id]["total_max"] += 100.0
                score_map[skill.id]["data_points"] += 1

    # Build profile items
    profile_items: list[SkillProfileItem] = []
    for skill in all_skills:
        data = score_map.get(skill.id)
        if data and data["data_points"] > 0:
            dp = data["data_points"]
            proficiency = (
                (data["total_score"] / data["total_max"] * 100)
                if data["total_max"] > 0 else None
            )
            if dp < 3:
                confidence = "insufficient"
                proficiency = None  # Not enough data to show
            elif dp <= 5:
                confidence = "low"
            elif dp <= 10:
                confidence = "moderate"
            else:
                confidence = "high"
        else:
            dp = 0
            proficiency = None
            confidence = "insufficient"

        profile_items.append(SkillProfileItem(
            skill_id=skill.id,
            skill_name=skill.name,
            skill_slug=skill.slug,
            skill_color=skill.color,
            skill_icon=skill.icon,
            proficiency=round(proficiency, 1) if proficiency is not None else None,
            data_points=dp,
            confidence=confidence,
            trend=trend_map.get(skill.id),
        ))

    return StudentSkillProfileResponse(
        student_id=student_id,
        student_name=student_name,
        skills=profile_items,
        total_ai_assignments_completed=total_ai_assignments,
    )


@router.get(
    "/me/skill-profile",
    response_model=StudentSkillProfileResponse,
    summary="Get own skill profile (Story 30.14)",
)
async def get_my_skill_profile(
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.student),
) -> StudentSkillProfileResponse:
    """
    Get skill profile for the authenticated student.

    Returns radar chart data with per-skill proficiency, confidence, and trend.
    """
    result = await session.execute(
        select(Student).where(Student.user_id == current_user.id)
    )
    student = result.scalar_one_or_none()

    if not student:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student record not found"
        )

    return await _build_skill_profile(
        student_id=student.id,
        student_name=current_user.full_name,
        session=session,
    )


@router.get(
    "/{student_id}/skill-profile",
    response_model=StudentSkillProfileResponse,
    summary="Get student skill profile (Story 30.14)",
)
async def get_student_skill_profile(
    student_id: uuid.UUID,
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.teacher, UserRole.admin, UserRole.supervisor),
) -> StudentSkillProfileResponse:
    """
    Get skill profile for a student.

    **Authorization:** Teacher must have access to student via their classes.
    Admin/Supervisor have full access.
    """
    # Verify student exists
    student_result = await session.execute(
        select(Student, User)
        .join(User, Student.user_id == User.id)
        .where(Student.id == student_id)
    )
    row = student_result.first()
    if not row:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Student not found"
        )
    student, student_user = row

    # Teacher access check
    if current_user.role == UserRole.teacher:
        teacher_result = await session.execute(
            select(Teacher).where(Teacher.user_id == current_user.id)
        )
        teacher = teacher_result.scalar_one_or_none()

        if not teacher:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Teacher record not found"
            )

        access_check = await session.execute(
            select(ClassStudent.id)
            .join(Class, ClassStudent.class_id == Class.id)
            .where(
                Class.teacher_id == teacher.id,
                ClassStudent.student_id == student_id,
            )
            .limit(1)
        )
        if access_check.scalar_one_or_none() is None:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="You do not have access to this student's data"
            )

    return await _build_skill_profile(
        student_id=student.id,
        student_name=student_user.full_name,
        session=session,
    )


# =============================================================================
# Student Skill Trend Endpoints (Story 30.16)
# =============================================================================


async def _build_skill_trends(
    student_id: uuid.UUID,
    period: str,
    session: AsyncSessionDep,
) -> StudentSkillTrendsResponse:
    """Build skill trend lines for a student, with fallback to
    deriving from AssignmentStudent scores when no StudentSkillScore
    records exist."""
    from datetime import timedelta

    # Calculate start date from period
    now = datetime.now(UTC)
    if period == "30d":
        start_date = now - timedelta(days=30)
    elif period == "3m":
        start_date = now - timedelta(days=90)
    elif period == "semester":
        # Academic semester: Fall=Sep-Jan, Spring=Feb-Jun
        month = now.month
        if 2 <= month <= 6:
            start_date = now.replace(month=2, day=1, hour=0, minute=0, second=0)
        elif 9 <= month <= 12:
            start_date = now.replace(month=9, day=1, hour=0, minute=0, second=0)
        else:  # Jan
            start_date = now.replace(year=now.year - 1, month=9, day=1, hour=0, minute=0, second=0)
    else:  # "all"
        start_date = datetime(2000, 1, 1, tzinfo=UTC)

    # Get all active skills
    skills_result = await session.execute(
        select(SkillCategory).where(SkillCategory.is_active == True)
    )
    all_skills = {s.id: s for s in skills_result.scalars().all()}
    skill_by_slug: dict[str, SkillCategory] = {s.slug: s for s in all_skills.values()}

    # Query scores with assignment names from StudentSkillScore
    query = (
        select(
            StudentSkillScore.skill_id,
            StudentSkillScore.attributed_score,
            StudentSkillScore.attributed_max_score,
            StudentSkillScore.recorded_at,
            StudentSkillScore.cefr_level,
            Assignment.name.label("assignment_name"),
        )
        .join(Assignment, StudentSkillScore.assignment_id == Assignment.id)
        .where(
            StudentSkillScore.student_id == student_id,
            StudentSkillScore.recorded_at >= start_date,
        )
        .order_by(StudentSkillScore.recorded_at.asc())
    )
    result = await session.execute(query)
    rows = result.all()

    # Group by skill
    skill_points: dict[uuid.UUID, list[SkillTrendPoint]] = {
        sid: [] for sid in all_skills
    }

    if rows:
        # Primary path: use StudentSkillScore records
        for row in rows:
            if row.skill_id not in all_skills:
                continue
            score = (
                (row.attributed_score / row.attributed_max_score * 100)
                if row.attributed_max_score > 0 else 0
            )
            skill_points[row.skill_id].append(SkillTrendPoint(
                date=row.recorded_at.strftime("%Y-%m-%d"),
                score=round(score, 1),
                assignment_name=row.assignment_name,
                cefr_level=row.cefr_level,
            ))
    else:
        # Fallback: derive from completed AI content assignments directly
        fallback_query = await session.execute(
            select(
                AssignmentStudent.score,
                AssignmentStudent.completed_at,
                Assignment.activity_type,
                Assignment.name,
            )
            .join(Assignment, AssignmentStudent.assignment_id == Assignment.id)
            .where(
                AssignmentStudent.student_id == student_id,
                AssignmentStudent.status == AssignmentStatus.completed,
                AssignmentStudent.score.isnot(None),
                AssignmentStudent.completed_at.isnot(None),
                AssignmentStudent.completed_at >= start_date,
                Assignment.activity_type.isnot(None),
            )
            .order_by(AssignmentStudent.completed_at.asc())
        )
        for row in fallback_query.all():
            act_type_val = row.activity_type.value if hasattr(row.activity_type, 'value') else str(row.activity_type)
            skill_slug = _ACTIVITY_TYPE_SKILL_SLUG.get(act_type_val)
            if not skill_slug:
                continue
            skill = skill_by_slug.get(skill_slug)
            if not skill:
                continue
            skill_points[skill.id].append(SkillTrendPoint(
                date=row.completed_at.strftime("%Y-%m-%d"),
                score=round(float(row.score), 1),
                assignment_name=row.name,
                cefr_level=None,
            ))

    # Build trend lines
    trends: list[SkillTrendLine] = []
    for skill_id, skill in all_skills.items():
        points = skill_points[skill_id]
        trends.append(SkillTrendLine(
            skill_id=skill_id,
            skill_name=skill.name,
            skill_slug=skill.slug,
            skill_color=skill.color,
            data_points=points,
            has_sufficient_data=len(points) >= 3,
        ))

    return StudentSkillTrendsResponse(
        student_id=student_id,
        period=period,
        trends=trends,
    )


@router.get(
    "/me/skill-trends",
    response_model=StudentSkillTrendsResponse,
    summary="Get own skill trends (Story 30.16)",
)
async def get_my_skill_trends(
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.student),
    period: str = Query(default="3m", description="Time period: 30d, 3m, semester, all"),
) -> StudentSkillTrendsResponse:
    """Get skill trend lines for the authenticated student."""
    result = await session.execute(
        select(Student).where(Student.user_id == current_user.id)
    )
    student = result.scalar_one_or_none()
    if not student:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student record not found")

    return await _build_skill_trends(student.id, period, session)


@router.get(
    "/{student_id}/skill-trends",
    response_model=StudentSkillTrendsResponse,
    summary="Get student skill trends (Story 30.16)",
)
async def get_student_skill_trends(
    student_id: uuid.UUID,
    session: AsyncSessionDep,
    current_user: User = require_role(UserRole.teacher, UserRole.admin, UserRole.supervisor),
    period: str = Query(default="3m", description="Time period: 30d, 3m, semester, all"),
) -> StudentSkillTrendsResponse:
    """Get skill trend lines for a student. Teacher must have access via classes."""
    # Verify student exists
    student_result = await session.execute(
        select(Student).where(Student.id == student_id)
    )
    if not student_result.scalar_one_or_none():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Student not found")

    # Teacher access check
    if current_user.role == UserRole.teacher:
        teacher_result = await session.execute(
            select(Teacher).where(Teacher.user_id == current_user.id)
        )
        teacher = teacher_result.scalar_one_or_none()
        if not teacher:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Teacher record not found")

        access_check = await session.execute(
            select(ClassStudent.id)
            .join(Class, ClassStudent.class_id == Class.id)
            .where(Class.teacher_id == teacher.id, ClassStudent.student_id == student_id)
            .limit(1)
        )
        if access_check.scalar_one_or_none() is None:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="You do not have access to this student's data")

    return await _build_skill_trends(student_id, period, session)
