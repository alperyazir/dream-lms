"""Excel report generator using openpyxl - Story 5.6."""

from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Template type display names
TEMPLATE_NAMES = {
    "weekly_class_summary": "Weekly Class Summary",
    "student_progress_report": "Student Progress Report",
    "monthly_assignment_overview": "Monthly Assignment Overview",
    "parent_teacher_conference": "Parent-Teacher Conference Report",
}

# Report type display names
REPORT_TYPE_NAMES = {
    "student": "Student Report",
    "class": "Class Report",
    "assignment": "Assignment Overview",
}

# Color scheme
HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SUBHEADER_FILL = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
ALT_ROW_FILL = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
BORDER = Border(
    left=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
    top=Side(style="thin", color="CBD5E0"),
    bottom=Side(style="thin", color="CBD5E0"),
)
TITLE_FONT = Font(bold=True, size=14, color="1A365D")
SECTION_FONT = Font(bold=True, size=12, color="2C5282")


def generate_excel_report(
    data: dict,
    report_type: str,
    template_type: str | None,
    output_path: str,
) -> None:
    """
    Generate an Excel report from report data.

    Args:
        data: Report data dictionary
        report_type: Type of report (student, class, assignment)
        template_type: Optional template type
        output_path: Path to write Excel file
    """
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create Summary sheet
    _create_summary_sheet(wb, data, report_type, template_type)

    # Create Data sheet based on report type
    if report_type == "student":
        _create_student_data_sheet(wb, data)
    elif report_type == "class":
        _create_class_data_sheet(wb, data)
    else:  # assignment
        _create_assignment_data_sheet(wb, data)

    # Create Charts sheet if applicable
    _create_charts_sheet(wb, data, report_type)

    # Save workbook
    wb.save(output_path)


def _create_summary_sheet(
    wb: Workbook, data: dict, report_type: str, template_type: str | None
) -> None:
    """Create the summary sheet."""
    ws = wb.create_sheet("Summary")

    # Title
    if template_type and template_type in TEMPLATE_NAMES:
        title = TEMPLATE_NAMES[template_type]
    else:
        title = REPORT_TYPE_NAMES.get(report_type, "Report")

    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # Report info
    row = 3
    ws[f"A{row}"] = "Report Generated:"
    ws[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    row += 1
    ws[f"A{row}"] = "Period:"
    ws[f"B{row}"] = f"{data.get('period_start', '')} to {data.get('period_end', '')}"

    row += 1
    if report_type == "student":
        ws[f"A{row}"] = "Student:"
        ws[f"B{row}"] = data.get("student_name", "Unknown")
        row += 1
        ws[f"A{row}"] = "Class:"
        ws[f"B{row}"] = data.get("class_name", "N/A")
    elif report_type == "class":
        ws[f"A{row}"] = "Class:"
        ws[f"B{row}"] = data.get("class_name", "Unknown")
        row += 1
        ws[f"A{row}"] = "Teacher:"
        ws[f"B{row}"] = data.get("teacher_name", "N/A")
        row += 1
        ws[f"A{row}"] = "Students:"
        ws[f"B{row}"] = data.get("student_count", 0)
    else:
        ws[f"A{row}"] = "Teacher:"
        ws[f"B{row}"] = data.get("teacher_name", "Unknown")
        row += 1
        ws[f"A{row}"] = "Total Assignments:"
        ws[f"B{row}"] = data.get("total_assignments", 0)

    # Summary Statistics section
    row += 2
    ws[f"A{row}"] = "Summary Statistics"
    ws[f"A{row}"].font = SECTION_FONT

    summary = data.get("summary", {})
    if isinstance(summary, dict):
        row += 1
        headers = ["Metric", "Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        metrics = [
            ("Average Score", f"{summary.get('avg_score', 0):.1f}%"),
            ("Total Completed", summary.get("total_completed", 0)),
            ("Completion Rate", f"{summary.get('completion_rate', 0) * 100:.0f}%"),
            ("Total Assigned", summary.get("total_assigned", 0)),
        ]

        for metric, value in metrics:
            row += 1
            ws.cell(row=row, column=1, value=metric).border = BORDER
            ws.cell(row=row, column=2, value=value).border = BORDER

    # Trend Analysis section
    trend = data.get("trend", {})
    if trend:
        row += 2
        ws[f"A{row}"] = "Trend Analysis"
        ws[f"A{row}"].font = SECTION_FONT

        row += 1
        direction = trend.get("direction", "new")
        change = trend.get("change")
        current = trend.get("current", 0)
        previous = trend.get("previous")

        ws.cell(row=row, column=1, value="Direction").border = BORDER
        ws.cell(row=row, column=2, value=direction.capitalize()).border = BORDER
        row += 1
        ws.cell(row=row, column=1, value="Current Period Avg").border = BORDER
        ws.cell(row=row, column=2, value=f"{current:.1f}%").border = BORDER
        row += 1
        ws.cell(row=row, column=1, value="Previous Period Avg").border = BORDER
        ws.cell(
            row=row, column=2, value=f"{previous:.1f}%" if previous else "N/A"
        ).border = BORDER
        row += 1
        ws.cell(row=row, column=1, value="Change").border = BORDER
        ws.cell(
            row=row, column=2, value=f"{change:+.1f}%" if change else "N/A"
        ).border = BORDER

    # Narrative section
    narrative = data.get("narrative", "")
    if narrative:
        row += 2
        ws[f"A{row}"] = "Summary"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1
        ws[f"A{row}"] = narrative
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20


def _create_student_data_sheet(wb: Workbook, data: dict) -> None:
    """Create student data detail sheet."""
    ws = wb.create_sheet("Details")

    row = 1

    # Skill Breakdown
    skill_breakdown = data.get("skill_breakdown", [])
    if skill_breakdown:
        ws[f"A{row}"] = "Skill Performance"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Skill", "Average Score", "Activities"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, item in enumerate(skill_breakdown):
            row += 1
            ws.cell(
                row=row,
                column=1,
                value=item.get("skill_name", "Unknown"),
            ).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{item.get('avg_score', 0):.1f}%"
            ).border = BORDER
            ws.cell(row=row, column=3, value=item.get("count", 0)).border = BORDER
            if i % 2 == 1:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        row += 2

    # Assignment List
    assignments = data.get("assignments", [])
    if assignments:
        ws[f"A{row}"] = "Assignment Details"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Assignment", "Score", "Time Spent", "Completed"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, item in enumerate(assignments):
            row += 1
            ws.cell(row=row, column=1, value=item.get("name", "Unknown")).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{item.get('score', 0)}%"
            ).border = BORDER
            ws.cell(
                row=row, column=3, value=f"{item.get('time_spent', 0)} min"
            ).border = BORDER
            completed = item.get("completed_at", "")
            ws.cell(
                row=row, column=4, value=completed[:10] if completed else "N/A"
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 5):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

    # Adjust column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15


def _create_class_data_sheet(wb: Workbook, data: dict) -> None:
    """Create class data detail sheet."""
    ws = wb.create_sheet("Details")

    row = 1

    # Score Distribution
    score_distribution = data.get("score_distribution", [])
    if score_distribution:
        ws[f"A{row}"] = "Score Distribution"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Score Range", "Number of Students"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, bucket in enumerate(score_distribution):
            row += 1
            ws.cell(
                row=row, column=1, value=bucket.get("range_label", "Unknown")
            ).border = BORDER
            ws.cell(row=row, column=2, value=bucket.get("count", 0)).border = BORDER
            if i % 2 == 1:
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        row += 2

    # Top Students
    top_students = data.get("top_students", [])
    if top_students:
        ws[f"A{row}"] = "Top Performing Students"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Rank", "Student", "Average Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, student in enumerate(top_students):
            row += 1
            ws.cell(row=row, column=1, value=student.get("rank", "")).border = BORDER
            ws.cell(
                row=row, column=2, value=student.get("name", "Unknown")
            ).border = BORDER
            ws.cell(
                row=row, column=3, value=f"{student.get('avg_score', 0):.1f}%"
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        row += 2

    # Struggling Students
    struggling_students = data.get("struggling_students", [])
    if struggling_students:
        ws[f"A{row}"] = "Students Needing Support"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Student", "Average Score", "Alert"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, student in enumerate(struggling_students):
            row += 1
            ws.cell(
                row=row, column=1, value=student.get("name", "Unknown")
            ).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{student.get('avg_score', 0):.1f}%"
            ).border = BORDER
            ws.cell(
                row=row, column=3, value=student.get("alert_reason", "")
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        row += 2

    # Assignment Performance
    assignments = data.get("assignments", [])
    if assignments:
        ws[f"A{row}"] = "Assignment Performance"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Assignment", "Average Score", "Completion Rate"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, item in enumerate(assignments):
            row += 1
            ws.cell(row=row, column=1, value=item.get("name", "Unknown")).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{item.get('avg_score', 0):.1f}%"
            ).border = BORDER
            ws.cell(
                row=row, column=3, value=f"{item.get('completion_rate', 0) * 100:.0f}%"
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 4):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

    # Adjust column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20


def _create_assignment_data_sheet(wb: Workbook, data: dict) -> None:
    """Create assignment overview data sheet."""
    ws = wb.create_sheet("Details")

    row = 1

    # Assignment Metrics
    assignments = data.get("assignments", [])
    if assignments:
        ws[f"A{row}"] = "Assignment Performance"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Assignment", "Avg Score", "Completion", "Avg Time", "Activity Type"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, item in enumerate(assignments):
            row += 1
            ws.cell(row=row, column=1, value=item.get("name", "Unknown")).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{item.get('avg_score', 0):.1f}%"
            ).border = BORDER
            ws.cell(
                row=row, column=3, value=f"{item.get('completion_rate', 0) * 100:.0f}%"
            ).border = BORDER
            ws.cell(
                row=row, column=4, value=f"{item.get('time_spent', 0):.0f} min"
            ).border = BORDER
            ws.cell(
                row=row, column=5, value=item.get("activity_type", "Unknown")
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 6):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        row += 2

    # Most Successful
    most_successful = data.get("most_successful", [])
    if most_successful:
        ws[f"A{row}"] = "Most Successful Assignments"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Assignment", "Average Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, item in enumerate(most_successful):
            row += 1
            ws.cell(row=row, column=1, value=item.get("name", "Unknown")).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{item.get('avg_score', 0):.1f}%"
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

        row += 2

    # Least Successful
    least_successful = data.get("least_successful", [])
    if least_successful:
        ws[f"A{row}"] = "Assignments Needing Attention"
        ws[f"A{row}"].font = SECTION_FONT
        row += 1

        headers = ["Assignment", "Average Score"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = BORDER

        for i, item in enumerate(least_successful):
            row += 1
            ws.cell(row=row, column=1, value=item.get("name", "Unknown")).border = BORDER
            ws.cell(
                row=row, column=2, value=f"{item.get('avg_score', 0):.1f}%"
            ).border = BORDER
            if i % 2 == 1:
                for col in range(1, 3):
                    ws.cell(row=row, column=col).fill = ALT_ROW_FILL

    # Adjust column widths
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 20


def _create_charts_sheet(wb: Workbook, data: dict, report_type: str) -> None:
    """Create charts sheet with Excel native charts."""
    # Use skill breakdown for student/class, fallback to activity_type_comparison for assignment
    if report_type in ("student", "class"):
        chart_data = data.get("skill_breakdown", [])
        chart_title = "Average Score by Skill"
        x_title = "Skill"
        label_key = "skill_name"
    else:
        chart_data = data.get("activity_type_comparison", [])
        chart_title = "Average Score by Activity Type"
        x_title = "Activity Type"
        label_key = "label"

    if not chart_data:
        return

    ws = wb.create_sheet("Charts")

    # Set up data for book/activity chart
    ws["A1"] = chart_title
    ws["A1"].font = SECTION_FONT

    ws["A3"] = x_title
    ws["B3"] = "Average Score"
    ws["A3"].fill = HEADER_FILL
    ws["A3"].font = HEADER_FONT
    ws["B3"].fill = HEADER_FILL
    ws["B3"].font = HEADER_FONT

    for i, item in enumerate(chart_data, 4):
        label = item.get(label_key, item.get("activity_type", "Unknown"))
        score = item.get("avg_score", 0)
        ws[f"A{i}"] = label
        ws[f"B{i}"] = score

    # Create bar chart
    if len(chart_data) > 0:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = chart_title
        chart.y_axis.title = "Score (%)"
        chart.x_axis.title = x_title

        data_ref = Reference(ws, min_col=2, min_row=3, max_row=3 + len(chart_data))
        cats_ref = Reference(ws, min_col=1, min_row=4, max_row=3 + len(chart_data))

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.shape = 4
        chart.width = 15
        chart.height = 10

        ws.add_chart(chart, "D3")

    # Add score distribution chart for class reports
    if report_type == "class":
        score_dist = data.get("score_distribution", [])
        if score_dist:
            start_row = 4 + len(chart_data) + 3

            ws[f"A{start_row}"] = "Score Distribution"
            ws[f"A{start_row}"].font = SECTION_FONT

            ws[f"A{start_row + 2}"] = "Score Range"
            ws[f"B{start_row + 2}"] = "Students"
            ws[f"A{start_row + 2}"].fill = HEADER_FILL
            ws[f"A{start_row + 2}"].font = HEADER_FONT
            ws[f"B{start_row + 2}"].fill = HEADER_FILL
            ws[f"B{start_row + 2}"].font = HEADER_FONT

            for i, bucket in enumerate(score_dist):
                ws[f"A{start_row + 3 + i}"] = bucket.get("range_label", "")
                ws[f"B{start_row + 3 + i}"] = bucket.get("count", 0)

            chart2 = BarChart()
            chart2.type = "col"
            chart2.style = 12
            chart2.title = "Score Distribution"
            chart2.y_axis.title = "Number of Students"
            chart2.x_axis.title = "Score Range"

            data_ref2 = Reference(
                ws,
                min_col=2,
                min_row=start_row + 2,
                max_row=start_row + 2 + len(score_dist),
            )
            cats_ref2 = Reference(
                ws,
                min_col=1,
                min_row=start_row + 3,
                max_row=start_row + 2 + len(score_dist),
            )

            chart2.add_data(data_ref2, titles_from_data=True)
            chart2.set_categories(cats_ref2)
            chart2.width = 15
            chart2.height = 10

            ws.add_chart(chart2, f"D{start_row}")

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
