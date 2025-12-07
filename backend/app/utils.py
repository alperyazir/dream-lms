import io
import logging
import re
import secrets
import string
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

import emails  # type: ignore
import jwt
from fastapi import UploadFile
from jinja2 import Template
from jwt.exceptions import InvalidTokenError
from openpyxl import load_workbook
from sqlmodel import Session
from unidecode import unidecode

from app.core import security
from app.core.config import settings

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@dataclass
class EmailData:
    html_content: str
    subject: str


def render_email_template(*, template_name: str, context: dict[str, Any]) -> str:
    template_str = (
        Path(__file__).parent / "email-templates" / "build" / template_name
    ).read_text()
    html_content = Template(template_str).render(context)
    return html_content


def send_email(
    *,
    email_to: str,
    subject: str = "",
    html_content: str = "",
) -> None:
    assert settings.emails_enabled, "no provided configuration for email variables"
    message = emails.Message(
        subject=subject,
        html=html_content,
        mail_from=(settings.EMAILS_FROM_NAME, settings.EMAILS_FROM_EMAIL),
    )
    smtp_options = {"host": settings.SMTP_HOST, "port": settings.SMTP_PORT}
    if settings.SMTP_TLS:
        smtp_options["tls"] = True
    elif settings.SMTP_SSL:
        smtp_options["ssl"] = True
    if settings.SMTP_USER:
        smtp_options["user"] = settings.SMTP_USER
    if settings.SMTP_PASSWORD:
        smtp_options["password"] = settings.SMTP_PASSWORD
    response = message.send(to=email_to, smtp=smtp_options)
    logger.info(f"send email result: {response}")


def generate_test_email(email_to: str) -> EmailData:
    project_name = settings.PROJECT_NAME
    subject = f"{project_name} - Test email"
    html_content = render_email_template(
        template_name="test_email.html",
        context={"project_name": settings.PROJECT_NAME, "email": email_to},
    )
    return EmailData(html_content=html_content, subject=subject)


def generate_reset_password_email(email_to: str, email: str, token: str) -> EmailData:
    project_name = settings.PROJECT_NAME
    subject = f"{project_name} - Password recovery for user {email}"
    link = f"{settings.FRONTEND_HOST}/reset-password?token={token}"
    html_content = render_email_template(
        template_name="reset_password.html",
        context={
            "project_name": settings.PROJECT_NAME,
            "username": email,
            "email": email_to,
            "valid_hours": settings.EMAIL_RESET_TOKEN_EXPIRE_HOURS,
            "link": link,
        },
    )
    return EmailData(html_content=html_content, subject=subject)


def generate_new_account_email(
    email_to: str, username: str, password: str
) -> EmailData:
    project_name = settings.PROJECT_NAME
    subject = f"{project_name} - New account for user {username}"
    html_content = render_email_template(
        template_name="new_account.html",
        context={
            "project_name": settings.PROJECT_NAME,
            "username": username,
            "password": password,
            "email": email_to,
            "link": settings.FRONTEND_HOST,
        },
    )
    return EmailData(html_content=html_content, subject=subject)


def generate_password_reset_token(email: str) -> str:
    delta = timedelta(hours=settings.EMAIL_RESET_TOKEN_EXPIRE_HOURS)
    now = datetime.now(timezone.utc)
    expires = now + delta
    exp = expires.timestamp()
    encoded_jwt = jwt.encode(
        {"exp": exp, "nbf": now, "sub": email},
        settings.SECRET_KEY,
        algorithm=security.ALGORITHM,
    )
    return encoded_jwt


def verify_password_reset_token(token: str) -> str | None:
    try:
        decoded_token = jwt.decode(
            token, settings.SECRET_KEY, algorithms=[security.ALGORITHM]
        )
        return str(decoded_token["sub"])
    except InvalidTokenError:
        return None


def generate_temp_password(length: int = 12) -> str:
    """
    Generate a cryptographically secure random temporary password.

    Args:
        length: Length of password to generate (default: 12)

    Returns:
        Random password string containing uppercase, lowercase, digits, and special characters

    Example:
        >>> password = generate_temp_password()
        >>> len(password)
        12
    """
    alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
    return ''.join(secrets.choice(alphabet) for _ in range(length))


# Turkish character mapping for username generation (Story 9.9)
TURKISH_CHAR_MAP: dict[str, str] = {
    'ı': 'i', 'İ': 'I',
    'ğ': 'g', 'Ğ': 'G',
    'ü': 'u', 'Ü': 'U',
    'ş': 's', 'Ş': 'S',
    'ö': 'o', 'Ö': 'O',
    'ç': 'c', 'Ç': 'C',
}


def turkish_to_ascii(text: str) -> str:
    """
    Convert Turkish special characters to ASCII equivalents.

    Args:
        text: Input text containing Turkish characters

    Returns:
        Text with Turkish characters replaced with ASCII equivalents

    Example:
        >>> turkish_to_ascii("Öğrenci")
        "Ogrenci"
    """
    result = []
    for char in text:
        result.append(TURKISH_CHAR_MAP.get(char, char))
    return ''.join(result)


def generate_username_from_fullname(full_name: str) -> str:
    """
    Generate a username from a full name for bulk import.

    Rules (Story 9.9 AC 22):
    - Converts to lowercase
    - Maps Turkish characters (ığüşöç → igusoc)
    - Replaces spaces with dots
    - Removes special characters except dots

    Args:
        full_name: User's full name

    Returns:
        Generated username (e.g., "ahmet.yilmaz")

    Example:
        >>> generate_username_from_fullname("Ahmet Yılmaz")
        "ahmet.yilmaz"
        >>> generate_username_from_fullname("Ömer Faruk Şahin")
        "omer.faruk.sahin"
    """
    if not full_name or not full_name.strip():
        return ""

    # Trim and apply Turkish character mapping
    name = full_name.strip()
    name = turkish_to_ascii(name)

    # Convert to lowercase
    name = name.lower()

    # Replace multiple spaces with single space
    name = re.sub(r'\s+', ' ', name)

    # Replace spaces with dots
    name = name.replace(' ', '.')

    # Remove all characters except alphanumeric and dots
    name = re.sub(r'[^a-z0-9.]', '', name)

    # Remove consecutive dots
    name = re.sub(r'\.+', '.', name)

    # Remove leading/trailing dots
    name = name.strip('.')

    return name


def ensure_unique_username(base_username: str, existing_usernames: set[str]) -> str:
    """
    Ensure username is unique by appending numbers if needed.

    Args:
        base_username: The base username to make unique
        existing_usernames: Set of already-used usernames

    Returns:
        Unique username (e.g., "john.doe", "john.doe2", "john.doe3")

    Example:
        >>> ensure_unique_username("john.doe", {"john.doe", "john.doe2"})
        "john.doe3"
    """
    if base_username not in existing_usernames:
        return base_username

    counter = 2
    while True:
        candidate = f"{base_username}{counter}"
        if candidate not in existing_usernames:
            return candidate
        counter += 1
        if counter > 1000:  # Safety limit
            raise ValueError(f"Could not generate unique username for '{base_username}'")


def generate_student_password(length: int = 8) -> str:
    """
    Generate an easy-to-type password for students.

    Rules (Story 9.9 AC 23):
    - 8 characters by default
    - Mix of letters and numbers
    - No confusing characters (0/O, 1/l/I)
    - Easy to type and communicate verbally

    Args:
        length: Length of password (default: 8)

    Returns:
        Generated password

    Example:
        >>> password = generate_student_password()
        >>> len(password)
        8
    """
    # Unambiguous characters - excluding 0, O, 1, l, I
    lowercase = 'abcdefghjkmnpqrstuvwxyz'  # no l
    uppercase = 'ABCDEFGHJKMNPQRSTUVWXYZ'  # no I, O
    digits = '23456789'  # no 0, 1

    # Ensure at least one of each type
    password_chars = [
        secrets.choice(lowercase),
        secrets.choice(uppercase),
        secrets.choice(digits),
    ]

    # Fill remaining with mixed characters
    all_chars = lowercase + uppercase + digits
    for _ in range(length - 3):
        password_chars.append(secrets.choice(all_chars))

    # Shuffle to randomize position
    secrets.SystemRandom().shuffle(password_chars)

    return ''.join(password_chars)


def generate_username(full_name: str, session: Session) -> str:
    """
    Generate unique username from full name.

    Algorithm:
    1. Extract first and last name from full_name
    2. Generate base: {first_initial}{last_name} (e.g., "John Doe" → "jdoe")
    3. Convert to lowercase, strip special characters
    4. Check uniqueness with crud.get_user_by_username()
    5. If exists, append incrementing number: jdoe1, jdoe2, etc.

    Args:
        full_name: User's full name (e.g., "John Doe")
        session: Database session for uniqueness checks

    Returns:
        Unique username (e.g., "jdoe", "jdoe1", etc.)

    Raises:
        ValueError: If full_name is empty or all variations taken

    Examples:
        >>> generate_username("John Doe", session)
        "jdoe"
        >>> generate_username("Madonna", session)
        "madonna"
        >>> generate_username("José García", session)
        "jgarcia"
    """
    # Import here to avoid circular dependency
    from app import crud

    if not full_name or not full_name.strip():
        raise ValueError("Full name cannot be empty")

    # Normalize unicode characters (José → Jose)
    normalized_name = unidecode(full_name.strip())

    # Remove special characters, keep only alphanumeric and spaces
    cleaned_name = re.sub(r'[^a-zA-Z0-9\s]', '', normalized_name)

    # Split into parts
    parts = cleaned_name.split()

    if not parts:
        raise ValueError("Full name must contain at least one valid character")

    # Generate base username
    if len(parts) == 1:
        # Single name: "Madonna" → "madonna"
        base_username = parts[0].lower()
    else:
        # Multiple parts: "John Doe" → "jdoe"
        first_initial = parts[0][0].lower()
        last_name = parts[-1].lower()
        base_username = f"{first_initial}{last_name}"

    # Validate base username format
    if not re.match(r'^[a-zA-Z0-9_-]+$', base_username):
        raise ValueError(f"Generated username '{base_username}' contains invalid characters")

    # Check uniqueness and increment if needed
    max_attempts = 100
    username = base_username

    for attempt in range(max_attempts):
        existing_user = crud.get_user_by_username(session=session, username=username)

        if not existing_user:
            return username

        # Username exists, try next variation
        username = f"{base_username}{attempt + 1}"

    # All variations taken
    raise ValueError(f"Could not generate unique username for '{full_name}' after {max_attempts} attempts")


async def validate_file_size(file: UploadFile, max_size_mb: int = 5) -> bool:
    """
    Validate file size does not exceed limit.

    Args:
        file: File to validate
        max_size_mb: Maximum file size in MB (default: 5)

    Returns:
        True if file size is within limit, False otherwise
    """
    contents = await file.read()
    await file.seek(0)  # Reset file pointer for later reads
    size_mb = len(contents) / (1024 * 1024)
    return size_mb <= max_size_mb


def validate_excel_headers(headers: list[str], required_headers: list[str]) -> bool:
    """
    Verify column names match expected headers.

    Args:
        headers: List of header names from Excel file
        required_headers: List of required header names

    Returns:
        True if all required headers are present, False otherwise
    """
    # Normalize headers by stripping whitespace and converting to lowercase
    normalized_headers = [h.strip().lower() if h else "" for h in headers]
    normalized_required = [h.strip().lower() for h in required_headers]

    return all(req in normalized_headers for req in normalized_required)


async def parse_excel_file(file: UploadFile) -> list[dict[str, Any]]:
    """
    Parse Excel file and return list of row dictionaries.

    Args:
        file: Excel file to parse

    Returns:
        List of dictionaries where keys are column headers and values are cell values.
        Each row dict includes '_row_number' for error reporting.

    Raises:
        ValueError: If file is corrupted, missing sheets, or cannot be read
    """
    try:
        contents = await file.read()
        await file.seek(0)  # Reset file pointer

        workbook = load_workbook(io.BytesIO(contents), read_only=True)

        if not workbook.worksheets:
            raise ValueError("Excel file contains no worksheets")

        sheet = workbook.active

        # First row contains headers
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)

        if not headers:
            raise ValueError("Excel file is empty or has no header row")

        # Parse data rows
        rows = []
        for idx, row in enumerate(rows_iter, start=1):  # Start at 1 for user-friendly display
            row_dict: dict[str, Any] = {}

            # Map headers to values
            for header, value in zip(headers, row, strict=False):
                if header:  # Skip empty header columns
                    row_dict[str(header)] = value

            # Add row number for error reporting (user-friendly, starting from 1)
            row_dict['_row_number'] = idx
            rows.append(row_dict)

        workbook.close()
        return rows

    except Exception as e:
        logger.error(f"Error parsing Excel file: {str(e)}")
        raise ValueError(f"Failed to parse Excel file: {str(e)}")
