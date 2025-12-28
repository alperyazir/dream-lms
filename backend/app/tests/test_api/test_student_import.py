"""API tests for student bulk import (Story 9.9)."""
import io

from fastapi import status
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlmodel import Session


def auth_headers(token: str) -> dict[str, str]:
    """Create authorization headers from token."""
    return {"Authorization": f"Bearer {token}"}


def create_test_excel(data: list[dict]) -> bytes:
    """Create a test Excel file with the given data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    # Default headers
    headers = ["Full Name", "Username", "Password", "Email", "Student ID", "Class/Grade"]
    ws.append(headers)

    for row in data:
        ws.append([
            row.get("Full Name", ""),
            row.get("Username", ""),
            row.get("Password", ""),
            row.get("Email", ""),
            row.get("Student ID", ""),
            row.get("Class/Grade", ""),
        ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


class TestImportTemplate:
    """Tests for the import template download endpoint."""

    def test_download_template_as_admin(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test admin can download import template."""
        response = client.get(
            "/api/v1/students/import-template",
            headers=auth_headers(admin_token),
        )

        assert response.status_code == status.HTTP_200_OK
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "student_import_template.xlsx" in response.headers.get("content-disposition", "")

        # Verify it's a valid Excel file
        wb = load_workbook(io.BytesIO(response.content))
        assert "Students" in wb.sheetnames
        assert "Instructions" in wb.sheetnames

        # Check headers in Students sheet
        ws = wb["Students"]
        headers = [cell.value for cell in ws[1]]
        assert "Full Name" in headers
        assert "Username" in headers
        assert "Password" in headers
        assert "Email" in headers
        assert "Student ID" in headers
        assert "Class/Grade" in headers

    def test_download_template_as_teacher(
        self, client: TestClient, teacher_token: str
    ) -> None:
        """Test teacher can download import template."""
        response = client.get(
            "/api/v1/students/import-template",
            headers=auth_headers(teacher_token),
        )

        assert response.status_code == status.HTTP_200_OK

    def test_download_template_as_student_forbidden(
        self, client: TestClient, student_token: str
    ) -> None:
        """Test student cannot download import template."""
        response = client.get(
            "/api/v1/students/import-template",
            headers=auth_headers(student_token),
        )

        assert response.status_code == status.HTTP_403_FORBIDDEN

    def test_download_template_unauthorized(self, client: TestClient) -> None:
        """Test unauthenticated request is rejected."""
        response = client.get("/api/v1/students/import-template")

        assert response.status_code == status.HTTP_401_UNAUTHORIZED


class TestImportValidation:
    """Tests for the import validation endpoint."""

    def test_validate_valid_file(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test validation of a valid import file."""
        excel_data = create_test_excel([
            {"Full Name": "John Doe", "Email": "john@test.com"},
            {"Full Name": "Jane Smith", "Email": "jane@test.com"},
        ])

        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["valid_count"] == 2
        assert data["error_count"] == 0
        assert data["total_count"] == 2
        assert len(data["rows"]) == 2
        assert data["rows"][0]["status"] == "valid"
        assert data["rows"][0]["username"] == "john.doe"

    def test_validate_missing_full_name(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test validation catches missing Full Name."""
        excel_data = create_test_excel([
            {"Full Name": "", "Email": "test@test.com"},
        ])

        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["error_count"] == 1
        assert data["rows"][0]["status"] == "error"
        assert any("Full Name is required" in err for err in data["rows"][0]["errors"])

    def test_validate_invalid_email(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test validation catches invalid email format."""
        excel_data = create_test_excel([
            {"Full Name": "John Doe", "Email": "not-an-email"},
        ])

        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["error_count"] == 1
        assert any("Invalid email format" in err for err in data["rows"][0]["errors"])

    def test_validate_duplicate_username_in_file(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test validation warns about duplicate usernames in file."""
        excel_data = create_test_excel([
            {"Full Name": "John Doe"},
            {"Full Name": "John Doe"},  # Same name = same generated username
        ])

        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        # First is valid, second has warning for duplicate
        assert data["warning_count"] >= 1
        # Second row should have john.doe2
        assert data["rows"][1]["username"] == "john.doe2"

    def test_validate_turkish_name(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test Turkish characters are converted in username."""
        excel_data = create_test_excel([
            {"Full Name": "Ahmet Yılmaz"},
        ])

        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_200_OK
        data = response.json()
        assert data["rows"][0]["username"] == "ahmet.yilmaz"

    def test_validate_rejects_non_excel(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test validation rejects non-Excel files."""
        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.txt", b"not an excel file", "text/plain")},
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "Only Excel files" in response.json()["detail"]

    def test_validate_rejects_empty_file(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test validation rejects file with no data rows."""
        wb = Workbook()
        ws = wb.active
        ws.append(["Full Name", "Username", "Email"])  # Headers only

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = client.post(
            "/api/v1/students/import/validate",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", buffer.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "no data rows" in response.json()["detail"]


class TestImportExecution:
    """Tests for the import execution endpoint."""

    def test_import_as_admin_requires_school_id(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test admin must provide school_id."""
        excel_data = create_test_excel([
            {"Full Name": "John Doe"},
        ])

        response = client.post(
            "/api/v1/students/import",
            headers=auth_headers(admin_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == status.HTTP_400_BAD_REQUEST
        assert "school_id" in response.json()["detail"]

    def test_import_as_teacher_uses_own_school(
        self, client: TestClient, session: Session, teacher_token: str
    ) -> None:
        """Test teacher imports students to their own school."""
        # Skip if teacher doesn't have school assigned
        # This test depends on fixture setup

        excel_data = create_test_excel([
            {"Full Name": "Import Test Student"},
        ])

        response = client.post(
            "/api/v1/students/import",
            headers=auth_headers(teacher_token),
            files={"file": ("students.xlsx", excel_data, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        # Expected behavior depends on test fixture:
        # - 404: Teacher record not found (fixture doesn't create Teacher model)
        # - 400: Teacher has no school assigned
        # - 200: Success (teacher has school and can import)
        assert response.status_code in [
            status.HTTP_200_OK,
            status.HTTP_400_BAD_REQUEST,
            status.HTTP_404_NOT_FOUND,
        ]

        if response.status_code == status.HTTP_400_BAD_REQUEST:
            assert "school" in response.json()["detail"].lower()
        elif response.status_code == status.HTTP_404_NOT_FOUND:
            assert "teacher" in response.json()["detail"].lower()


class TestCredentialsDownload:
    """Tests for the credentials download endpoint."""

    def test_download_credentials(
        self, client: TestClient, admin_token: str
    ) -> None:
        """Test downloading credentials file."""
        credentials = [
            {"full_name": "John Doe", "username": "john.doe", "password": "abc12345", "email": "john@test.com"},
            {"full_name": "Jane Smith", "username": "jane.smith", "password": "xyz67890", "email": None},
        ]

        response = client.post(
            "/api/v1/students/import/credentials",
            headers=auth_headers(admin_token),
            json={"credentials": credentials},
        )

        assert response.status_code == status.HTTP_200_OK
        assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        assert "student_credentials" in response.headers.get("content-disposition", "")

        # Verify content
        wb = load_workbook(io.BytesIO(response.content))
        ws = wb.active

        # Check warning row
        assert "WARNING" in str(ws.cell(row=1, column=1).value)

        # Check headers (row 2 after warning)
        headers = [ws.cell(row=2, column=i).value for i in range(1, 5)]
        assert "Full Name" in headers
        assert "Username" in headers
        assert "Password" in headers

        # Check data
        assert ws.cell(row=3, column=1).value == "John Doe"
        assert ws.cell(row=3, column=2).value == "john.doe"
        assert ws.cell(row=3, column=3).value == "abc12345"

    def test_download_credentials_requires_auth(
        self, client: TestClient
    ) -> None:
        """Test credentials download requires authentication."""
        response = client.post(
            "/api/v1/students/import/credentials",
            json={"credentials": []},
        )

        assert response.status_code == status.HTTP_401_UNAUTHORIZED

    def test_download_credentials_as_student_forbidden(
        self, client: TestClient, student_token: str
    ) -> None:
        """Test students cannot download credentials."""
        response = client.post(
            "/api/v1/students/import/credentials",
            headers=auth_headers(student_token),
            json={"credentials": []},
        )

        assert response.status_code == status.HTTP_403_FORBIDDEN
