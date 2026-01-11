"""
Tests for student password management - Story 28.1
Teacher-controlled student password management: view, set, and create with passwords.
"""
import uuid
from unittest.mock import patch

import pytest
from fastapi.testclient import TestClient
from sqlmodel import Session

from app.core.config import settings
from app.core.security import get_password_hash, verify_password
from app.models import (
    Class,
    ClassStudent,
    School,
    Student,
    Teacher,
    User,
    UserRole,
)


class TestGetStudentPassword:
    """Tests for GET /admin/students/{student_id}/password endpoint [Story 28.1]"""

    def test_admin_can_get_student_password(
        self, client: TestClient, session: Session, admin_token: str
    ):
        """Test admin can retrieve student password."""
        # Create student with viewable password
        student_user = User(
            id=uuid.uuid4(),
            email="test_student@example.com",
            username="test_student_pwd",
            hashed_password=get_password_hash("student123"),
            viewable_password_encrypted=None,  # Will be set via mock
            role=UserRole.student,
            is_active=True,
            full_name="Test Student"
        )
        session.add(student_user)
        session.commit()

        student = Student(
            id=uuid.uuid4(),
            user_id=student_user.id,
        )
        session.add(student)
        session.commit()

        # Mock encryption for password viewing
        with patch("app.api.routes.admin.decrypt_viewable_password") as mock_decrypt:
            mock_decrypt.return_value = "student123"
            student_user.viewable_password_encrypted = "encrypted_value"
            session.add(student_user)
            session.commit()

            response = client.get(
                f"{settings.API_V1_STR}/admin/students/{student.id}/password",
                headers={"Authorization": f"Bearer {admin_token}"},
            )

            assert response.status_code == 200
            data = response.json()
            assert data["student_id"] == str(student.id)
            assert data["username"] == "test_student_pwd"
            assert data["full_name"] == "Test Student"
            assert data["password"] == "student123"

    def test_student_not_found_returns_404(
        self, client: TestClient, admin_token: str
    ):
        """Test 404 is returned for non-existent student."""
        fake_id = uuid.uuid4()
        response = client.get(
            f"{settings.API_V1_STR}/admin/students/{fake_id}/password",
            headers={"Authorization": f"Bearer {admin_token}"},
        )

        assert response.status_code == 404
        assert "Student not found" in response.json()["detail"]

    def test_teacher_can_get_own_student_password(
        self, client: TestClient, session: Session
    ):
        """Test teacher can get password for student in their class."""
        # Create school
        school = School(
            id=uuid.uuid4(),
            name="Test School",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        # Create teacher user and record
        teacher_user = User(
            id=uuid.uuid4(),
            email="teacher_pwd@example.com",
            username="teacher_pwd",
            hashed_password=get_password_hash("teacherpassword"),
            role=UserRole.teacher,
            is_active=True,
            full_name="Teacher PWD"
        )
        session.add(teacher_user)
        session.commit()

        teacher = Teacher(
            id=uuid.uuid4(),
            user_id=teacher_user.id,
            school_id=school.id,
        )
        session.add(teacher)
        session.commit()

        # Create class owned by teacher
        class_obj = Class(
            id=uuid.uuid4(),
            name="Test Class",
            school_id=school.id,
            teacher_id=teacher.id,
        )
        session.add(class_obj)
        session.commit()

        # Create student
        student_user = User(
            id=uuid.uuid4(),
            email="student_in_class@example.com",
            username="student_in_class",
            hashed_password=get_password_hash("student123"),
            role=UserRole.student,
            is_active=True,
            full_name="Student In Class"
        )
        session.add(student_user)
        session.commit()

        student = Student(
            id=uuid.uuid4(),
            user_id=student_user.id,
            created_by_teacher_id=teacher.id,
        )
        session.add(student)
        session.commit()

        # Enroll student in class
        enrollment = ClassStudent(
            id=uuid.uuid4(),
            class_id=class_obj.id,
            student_id=student.id,
        )
        session.add(enrollment)
        session.commit()

        # Get teacher token
        response = client.post(
            f"{settings.API_V1_STR}/login/access-token",
            data={"username": teacher_user.email, "password": "teacherpassword"}
        )
        assert response.status_code == 200
        teacher_token = response.json()["access_token"]

        # Mock encryption for password viewing
        with patch("app.api.routes.admin.decrypt_viewable_password") as mock_decrypt:
            mock_decrypt.return_value = "student123"
            student_user.viewable_password_encrypted = "encrypted_value"
            session.add(student_user)
            session.commit()

            response = client.get(
                f"{settings.API_V1_STR}/admin/students/{student.id}/password",
                headers={"Authorization": f"Bearer {teacher_token}"},
            )

            assert response.status_code == 200
            data = response.json()
            assert data["password"] == "student123"

    def test_teacher_cannot_get_other_student_password(
        self, client: TestClient, session: Session
    ):
        """Test teacher cannot get password for student not in their class."""
        # Create school
        school = School(
            id=uuid.uuid4(),
            name="Test School 2",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        # Create teacher user and record
        teacher_user = User(
            id=uuid.uuid4(),
            email="teacher_no_access@example.com",
            username="teacher_no_access",
            hashed_password=get_password_hash("teacherpassword"),
            role=UserRole.teacher,
            is_active=True,
            full_name="Teacher No Access"
        )
        session.add(teacher_user)
        session.commit()

        teacher = Teacher(
            id=uuid.uuid4(),
            user_id=teacher_user.id,
            school_id=school.id,
        )
        session.add(teacher)
        session.commit()

        # Create student NOT in teacher's class or created by teacher
        student_user = User(
            id=uuid.uuid4(),
            email="other_student@example.com",
            username="other_student",
            hashed_password=get_password_hash("student123"),
            role=UserRole.student,
            is_active=True,
            full_name="Other Student"
        )
        session.add(student_user)
        session.commit()

        student = Student(
            id=uuid.uuid4(),
            user_id=student_user.id,
            # Not created by this teacher
        )
        session.add(student)
        session.commit()

        # Get teacher token
        response = client.post(
            f"{settings.API_V1_STR}/login/access-token",
            data={"username": teacher_user.email, "password": "teacherpassword"}
        )
        assert response.status_code == 200
        teacher_token = response.json()["access_token"]

        response = client.get(
            f"{settings.API_V1_STR}/admin/students/{student.id}/password",
            headers={"Authorization": f"Bearer {teacher_token}"},
        )

        assert response.status_code == 403
        assert "do not have access" in response.json()["detail"].lower()

    def test_password_not_available_for_pre_feature_student(
        self, client: TestClient, session: Session, admin_token: str
    ):
        """Test message shown for students without viewable password (created before feature)."""
        # Create student without viewable password
        student_user = User(
            id=uuid.uuid4(),
            email="old_student@example.com",
            username="old_student",
            hashed_password=get_password_hash("oldpassword"),
            viewable_password_encrypted=None,  # Pre-feature student
            role=UserRole.student,
            is_active=True,
            full_name="Old Student"
        )
        session.add(student_user)
        session.commit()

        student = Student(
            id=uuid.uuid4(),
            user_id=student_user.id,
        )
        session.add(student)
        session.commit()

        response = client.get(
            f"{settings.API_V1_STR}/admin/students/{student.id}/password",
            headers={"Authorization": f"Bearer {admin_token}"},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["password"] is None
        assert "not available" in data["message"].lower()


class TestSetStudentPassword:
    """Tests for PUT /admin/students/{student_id}/password endpoint [Story 28.1]"""

    def test_admin_can_set_student_password(
        self, client: TestClient, session: Session, admin_token: str
    ):
        """Test admin can set new password for student."""
        # Create student
        student_user = User(
            id=uuid.uuid4(),
            email="student_set_pwd@example.com",
            username="student_set_pwd",
            hashed_password=get_password_hash("oldpassword"),
            role=UserRole.student,
            is_active=True,
            full_name="Student Set PWD"
        )
        session.add(student_user)
        session.commit()

        student = Student(
            id=uuid.uuid4(),
            user_id=student_user.id,
        )
        session.add(student)
        session.commit()

        # Set new password
        with patch("app.api.routes.admin.encrypt_viewable_password") as mock_encrypt:
            mock_encrypt.return_value = "encrypted_new_password"

            response = client.put(
                f"{settings.API_V1_STR}/admin/students/{student.id}/password",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={"password": "newpassword123"},
            )

            assert response.status_code == 200
            data = response.json()
            assert data["password"] == "newpassword123"
            assert "updated" in data["message"].lower()

            # Verify password was actually changed
            session.refresh(student_user)
            assert verify_password("newpassword123", student_user.hashed_password)

    def test_set_password_validates_length(
        self, client: TestClient, session: Session, admin_token: str
    ):
        """Test password must meet minimum length requirement."""
        # Create student
        student_user = User(
            id=uuid.uuid4(),
            email="student_short_pwd@example.com",
            username="student_short_pwd",
            hashed_password=get_password_hash("oldpassword"),
            role=UserRole.student,
            is_active=True,
            full_name="Student Short PWD"
        )
        session.add(student_user)
        session.commit()

        student = Student(
            id=uuid.uuid4(),
            user_id=student_user.id,
        )
        session.add(student)
        session.commit()

        # Try to set too short password
        response = client.put(
            f"{settings.API_V1_STR}/admin/students/{student.id}/password",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"password": "ab"},  # Too short
        )

        assert response.status_code == 422  # Validation error

    def test_student_not_found_returns_404(
        self, client: TestClient, admin_token: str
    ):
        """Test 404 is returned for non-existent student."""
        fake_id = uuid.uuid4()
        response = client.put(
            f"{settings.API_V1_STR}/admin/students/{fake_id}/password",
            headers={"Authorization": f"Bearer {admin_token}"},
            json={"password": "newpassword"},
        )

        assert response.status_code == 404


class TestStudentCannotChangePassword:
    """Tests verifying students cannot change their own passwords [Story 28.1]"""

    def test_student_cannot_update_password_me(
        self, client: TestClient, student_token: str
    ):
        """Test student is blocked from changing password via /me/password endpoint."""
        response = client.patch(
            f"{settings.API_V1_STR}/users/me/password",
            headers={"Authorization": f"Bearer {student_token}"},
            json={
                "current_password": "studentpassword",
                "new_password": "newpassword123"
            },
        )

        assert response.status_code == 403
        assert "cannot change password" in response.json()["detail"].lower()

    def test_student_cannot_change_initial_password(
        self, client: TestClient, student_token: str
    ):
        """Test student is blocked from changing password via /me/change-initial-password."""
        response = client.post(
            f"{settings.API_V1_STR}/users/me/change-initial-password",
            headers={"Authorization": f"Bearer {student_token}"},
            json={
                "current_password": "studentpassword",
                "new_password": "newpassword123"
            },
        )

        assert response.status_code == 403
        assert "cannot change password" in response.json()["detail"].lower()


class TestStudentCreationWithPassword:
    """Tests for student creation with optional password [Story 28.1]"""

    def test_create_student_with_custom_password(
        self, client: TestClient, session: Session, admin_token: str
    ):
        """Test admin can create student with custom password."""
        # Create school
        school = School(
            id=uuid.uuid4(),
            name="Test School Custom PWD",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        with patch("app.api.routes.admin.encrypt_viewable_password") as mock_encrypt:
            mock_encrypt.return_value = "encrypted_custom"

            response = client.post(
                f"{settings.API_V1_STR}/admin/students",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={
                    "username": "custom_pwd_student",
                    "full_name": "Custom Password Student",
                    "school_id": str(school.id),
                    "password": "mycustompassword",
                },
            )

            assert response.status_code == 201
            data = response.json()

            # Verify password is returned in response
            assert "temporary_password" in data
            assert data["temporary_password"] == "mycustompassword"

            # Verify must_change_password is False for students
            assert data["user"]["must_change_password"] is False

    def test_create_student_with_auto_generated_password(
        self, client: TestClient, session: Session, admin_token: str
    ):
        """Test admin can create student with auto-generated password when none provided."""
        # Create school
        school = School(
            id=uuid.uuid4(),
            name="Test School Auto PWD",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        with patch("app.api.routes.admin.encrypt_viewable_password") as mock_encrypt:
            mock_encrypt.return_value = "encrypted_auto"

            response = client.post(
                f"{settings.API_V1_STR}/admin/students",
                headers={"Authorization": f"Bearer {admin_token}"},
                json={
                    "username": "auto_pwd_student",
                    "full_name": "Auto Password Student",
                    "school_id": str(school.id),
                    # No password provided - should auto-generate
                },
            )

            assert response.status_code == 201
            data = response.json()

            # Verify auto-generated password is returned
            assert "temporary_password" in data
            assert len(data["temporary_password"]) >= 8  # Auto-generated passwords are 8 chars


class TestBulkImportWithPassword:
    """Tests for bulk import password options [Story 28.1]"""

    def test_import_with_class_password(
        self, client: TestClient, session: Session
    ):
        """Test bulk import with class_password parameter."""
        import io

        from openpyxl import Workbook

        # Create school and teacher
        school = School(
            id=uuid.uuid4(),
            name="Bulk Import School",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        teacher_user = User(
            id=uuid.uuid4(),
            email="bulk_teacher@example.com",
            username="bulk_teacher",
            hashed_password=get_password_hash("teacherpassword"),
            role=UserRole.teacher,
            is_active=True,
            full_name="Bulk Teacher"
        )
        session.add(teacher_user)
        session.commit()

        teacher = Teacher(
            id=uuid.uuid4(),
            user_id=teacher_user.id,
            school_id=school.id,
        )
        session.add(teacher)
        session.commit()

        # Get teacher token
        response = client.post(
            f"{settings.API_V1_STR}/login/access-token",
            data={"username": teacher_user.email, "password": "teacherpassword"}
        )
        teacher_token = response.json()["access_token"]

        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Full Name *"
        ws["B1"] = "Username"
        ws["C1"] = "Password"
        ws["D1"] = "Email"
        ws["A2"] = "Test Student 1"
        ws["B2"] = ""  # Auto-generate username
        ws["C2"] = ""  # No password in file - should use class_password

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Import with class_password
        response = client.post(
            f"{settings.API_V1_STR}/students/import?class_password=classwide123",
            headers={"Authorization": f"Bearer {teacher_token}"},
            files={"file": ("students.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["created_count"] == 1

        # Verify the password used was the class password
        assert len(data["credentials"]) == 1
        assert data["credentials"][0]["password"] == "classwide123"

    def test_import_with_password_from_file(
        self, client: TestClient, session: Session
    ):
        """Test bulk import uses password from Excel file when provided."""
        import io

        from openpyxl import Workbook

        # Create school and teacher
        school = School(
            id=uuid.uuid4(),
            name="Bulk Import School 2",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        teacher_user = User(
            id=uuid.uuid4(),
            email="bulk_teacher2@example.com",
            username="bulk_teacher2",
            hashed_password=get_password_hash("teacherpassword"),
            role=UserRole.teacher,
            is_active=True,
            full_name="Bulk Teacher 2"
        )
        session.add(teacher_user)
        session.commit()

        teacher = Teacher(
            id=uuid.uuid4(),
            user_id=teacher_user.id,
            school_id=school.id,
        )
        session.add(teacher)
        session.commit()

        # Get teacher token
        response = client.post(
            f"{settings.API_V1_STR}/login/access-token",
            data={"username": teacher_user.email, "password": "teacherpassword"}
        )
        teacher_token = response.json()["access_token"]

        # Create Excel file with password in file
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Full Name *"
        ws["B1"] = "Username"
        ws["C1"] = "Password"
        ws["D1"] = "Email"
        ws["A2"] = "Test Student File PWD"
        ws["B2"] = ""
        ws["C2"] = "filepwd123"  # Password from file

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Import without class_password - should use password from file
        response = client.post(
            f"{settings.API_V1_STR}/students/import",
            headers={"Authorization": f"Bearer {teacher_token}"},
            files={"file": ("students.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["created_count"] == 1

        # Verify the password used was from the file
        assert len(data["credentials"]) == 1
        assert data["credentials"][0]["password"] == "filepwd123"

    def test_import_file_password_takes_priority_over_class_password(
        self, client: TestClient, session: Session
    ):
        """Test password from file takes priority over class_password."""
        import io

        from openpyxl import Workbook

        # Create school and teacher
        school = School(
            id=uuid.uuid4(),
            name="Bulk Import School Priority",
            dcs_publisher_id=999
        )
        session.add(school)
        session.commit()

        teacher_user = User(
            id=uuid.uuid4(),
            email="bulk_teacher_priority@example.com",
            username="bulk_teacher_priority",
            hashed_password=get_password_hash("teacherpassword"),
            role=UserRole.teacher,
            is_active=True,
            full_name="Bulk Teacher Priority"
        )
        session.add(teacher_user)
        session.commit()

        teacher = Teacher(
            id=uuid.uuid4(),
            user_id=teacher_user.id,
            school_id=school.id,
        )
        session.add(teacher)
        session.commit()

        # Get teacher token
        response = client.post(
            f"{settings.API_V1_STR}/login/access-token",
            data={"username": teacher_user.email, "password": "teacherpassword"}
        )
        teacher_token = response.json()["access_token"]

        # Create Excel file with password
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Full Name *"
        ws["B1"] = "Username"
        ws["C1"] = "Password"
        ws["D1"] = "Email"
        ws["A2"] = "Priority Test Student"
        ws["B2"] = ""
        ws["C2"] = "filepriority"  # This should take priority

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Import WITH class_password - file password should still take priority
        response = client.post(
            f"{settings.API_V1_STR}/students/import?class_password=classpwd",
            headers={"Authorization": f"Bearer {teacher_token}"},
            files={"file": ("students.xlsx", buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        )

        assert response.status_code == 200
        data = response.json()
        assert data["created_count"] == 1

        # Verify file password took priority
        assert data["credentials"][0]["password"] == "filepriority"
